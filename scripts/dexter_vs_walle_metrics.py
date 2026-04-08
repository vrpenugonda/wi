"""
DEXter vs WALLE comparison metrics (no gold labels required).

Metric 1 implemented:
  - Coverage / completeness at L1-L4 for DEX vs WALLE

Input:
  - merged CSV output from scripts/join_dexter_with_walle_snowflake.py (DEX_* + WALLE_* columns)

Output:
  - Excel report with:
      - data: standardized, readable column names
      - metric_1_coverage: coverage % table

Run (example):
  uv run python scripts/dexter_vs_walle_metrics.py --in merged.csv --out report.xlsx
"""

from __future__ import annotations

import argparse
import json
import random
import time
from pathlib import Path
from urllib import request, error
from urllib.parse import urlencode
from typing import Any, Callable

import pandas as pd
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl import load_workbook


def _is_usable_label(v, generic: set[str]) -> bool:
    if v is None or pd.isna(v):
        return False
    s = str(v).strip()
    if s == "":
        return False
    if s.lower() in generic:
        return False
    return True


def _normalize_label_series(s: pd.Series) -> pd.Series:
    """Normalize labels for comparison (stringify, strip, lower)."""
    return (
        s.fillna("")
        .astype(str)
        .map(lambda x: x.strip())
        .map(lambda x: x.lower())
    )


def _entropy_from_counts(counts: pd.Series) -> float:
    """Shannon entropy (base-2) from a value_counts series."""
    total = float(counts.sum())
    if total <= 0:
        return 0.0
    p = counts.astype(float) / total
    p = p[p > 0]
    import math

    return float(-sum(float(pi) * math.log2(float(pi)) for pi in p))


def _top_share(s: pd.Series, top_n: int) -> float:
    """Share (%) covered by top_n labels."""
    if s.empty:
        return 0.0
    vc = s.value_counts(dropna=False)
    total = float(vc.sum())
    if total <= 0:
        return 0.0
    return float(vc.head(top_n).sum() / total * 100.0)


def _usable_mask(s: pd.Series, generic: set[str]) -> pd.Series:
    """Boolean mask for usable labels."""
    return s.apply(lambda v: _is_usable_label(v, generic))


def _to_month_bucket(s: pd.Series) -> pd.Series:
    """
    Convert a datetime-like series to YYYY-MM strings (month buckets).
    Missing/unparseable values become empty string.
    """
    dt = pd.to_datetime(s, errors="coerce", utc=True)
    out = dt.dt.strftime("%Y-%m")
    return out.fillna("")


def _build_path(l1, l2, l3, l4, generic: set[str]) -> str | None:
    parts: list[str] = []
    for x in (l1, l2, l3, l4):
        if x is None or pd.isna(x):
            continue
        s = str(x).strip()
        if s == "":
            continue
        if s.lower() in generic:
            continue
        parts.append(s)
    return " → ".join(parts) if parts else None


def _path_res_sim(path: str | None, text: str | None) -> float | None:
    # sklearn vectorizers require real strings; NaN/None will raise.
    if path is None or pd.isna(path):
        return None
    if text is None or pd.isna(text):
        return None
    path = str(path).strip()
    text = str(text).strip()
    if path == "" or text == "":
        return None
    try:
        from sklearn.feature_extraction.text import TfidfVectorizer
        from sklearn.metrics.pairwise import cosine_similarity
    except Exception:
        return None
    v = TfidfVectorizer().fit([path, text])
    m = v.transform([path, text])
    return float(cosine_similarity(m[0], m[1])[0][0])


def _append_blocks_after_write(xlsx_path: Path, blocks: dict[str, list[dict]]) -> None:
    """
    Append summary blocks after the data table in a sheet.
    blocks[sheet] = [{"title": str, "headers": [...], "rows": [[...], ...]}, ...]
    """
    wb = load_workbook(xlsx_path)
    for sheet, blks in blocks.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for b in blks:
            ws.append([])
            title = b.get("title")
            if title:
                ws.append([title])
            headers = b.get("headers") or []
            if headers:
                ws.append(list(headers))
            for r in b.get("rows") or []:
                ws.append(list(r))
    wb.save(xlsx_path)

def _is_trueish(v) -> bool:
    if v is None or pd.isna(v):
        return False
    if isinstance(v, bool):
        return bool(v)
    s = str(v).strip().lower()
    return s in {"true", "t", "1", "yes", "y"}


def _pct(mask: pd.Series) -> float:
    if mask is None or len(mask) == 0:
        return 0.0
    return float(mask.mean() * 100.0)


def _azure_openai_chat_json(
    *,
    endpoint: str,
    api_key: str,
    deployment: str,
    api_version: str,
    messages: list[dict],
    max_completion_tokens: int = 600,
    temperature: float | None = None,
    use_response_format_json: bool = True,
) -> dict:
    """
    Minimal Azure OpenAI Chat Completions call (JSON response).

    Uses env/args compatible with the rest of this repo:
      - AZURE_OPENAI_ENDPOINT
      - AZURE_OPENAI_API_KEY
      - AZURE_OPENAI_DEPLOYMENT_NAME
      - AZURE_OPENAI_API_VERSION
    """
    endpoint = endpoint.rstrip("/")
    qs = urlencode({"api-version": api_version})
    url = f"{endpoint}/openai/deployments/{deployment}/chat/completions?{qs}"
    # Auth header selection:
    # - In this repo's workflows/classifiers we often use an Azure AD access token (JWT) from client-credentials.
    #   Azure OpenAI expects that as: Authorization: Bearer <token>
    # - If a real Azure OpenAI API key is provided instead, use: api-key: <key>
    token = (api_key or "").strip()
    is_jwt = token.startswith("eyJ") and token.count(".") >= 2
    headers = {"Content-Type": "application/json"}
    if is_jwt:
        headers["Authorization"] = f"Bearer {token}"
    else:
        headers["api-key"] = token

    # Best-effort: mirror gateway headers used by the main classifiers.
    # These may be required by upstream API gateways for routing/telemetry.
    import os
    x_upstream_env = (os.getenv("X_UPSTREAM_ENV") or os.getenv("X-Upstream-Env") or "").strip()
    project_id = (os.getenv("PROJECT_ID") or os.getenv("projectId") or "").strip()
    if x_upstream_env:
        headers["X-Upstream-Env"] = x_upstream_env
        headers["X-Model-Usage-Type"] = x_upstream_env
        headers["modelUsageType"] = x_upstream_env
    if project_id:
        headers["projectId"] = project_id
    payload = {
        "messages": messages,
        # Newer models (and some Azure deployments) require max_completion_tokens instead of max_tokens.
        "max_completion_tokens": max_completion_tokens,
    }
    # Some deployments (reasoning models) restrict/ignore temperature; omit it to use model default.
    if temperature is not None:
        payload["temperature"] = float(temperature)
    # Some Azure OpenAI deployments/api-versions don't support response_format.
    # We'll optionally include it, and fall back (retry) at call-site if needed.
    if use_response_format_json:
        payload["response_format"] = {"type": "json_object"}
    data = json.dumps(payload).encode("utf-8")
    req = request.Request(
        url,
        data=data,
        headers=headers,
        method="POST",
    )
    try:
        with request.urlopen(req, timeout=120) as resp:
            raw = resp.read().decode("utf-8", errors="replace")
            return json.loads(raw)
    except error.HTTPError as e:
        body = ""
        try:
            body = e.read().decode("utf-8", errors="replace")
        except Exception:
            pass
        raise RuntimeError(
            json.dumps(
                {
                    "type": "azure_openai_http_error",
                    "status_code": int(getattr(e, "code", 0) or 0),
                    "headers": dict(getattr(e, "headers", {}) or {}),
                    "body": body[:2000],
                }
            )
        ) from e


def _extract_json_content(resp: dict) -> dict:
    """
    Extract JSON object from Azure OpenAI response.
    """
    try:
        content = resp["choices"][0]["message"]["content"]
    except Exception as e:
        raise RuntimeError(f"Unexpected response shape: keys={list(resp.keys())}") from e
    if isinstance(content, dict):
        return content
    if not isinstance(content, str):
        raise RuntimeError(f"Unexpected message content type: {type(content)}")
    content = content.strip()
    # Some models wrap JSON in text; try to recover the first JSON object.
    if content.startswith("{") and content.endswith("}"):
        return json.loads(content)
    start = content.find("{")
    end = content.rfind("}")
    if start != -1 and end != -1 and end > start:
        return json.loads(content[start : end + 1])
    raise RuntimeError(f"Could not parse JSON from content: {content[:200]}")


def _maybe_sleep_for_rpm(last_call_ts: float | None, max_rpm: int, jitter_s: float = 0.25) -> float:
    """
    Proactive throttle to avoid exceeding service-side RPM limits.
    Returns the timestamp of "now" after any sleep.
    """
    if max_rpm <= 0:
        return time.time()
    min_interval = 60.0 / float(max_rpm)
    now = time.time()
    if last_call_ts is None:
        return now
    elapsed = now - last_call_ts
    remaining = min_interval - elapsed
    if remaining > 0:
        sleep_for = remaining + (random.random() * jitter_s)
        time.sleep(sleep_for)
    return time.time()


def _is_retryable_azure_error(e: Exception) -> tuple[bool, int | None, float | None]:
    """
    Detect retryable Azure OpenAI errors from our RuntimeError JSON.
    Returns (retryable, status_code, retry_after_seconds).
    """
    msg = str(e)
    try:
        obj = json.loads(msg)
        if obj.get("type") != "azure_openai_http_error":
            return False, None, None
        status = int(obj.get("status_code") or 0)
        headers = {str(k).lower(): str(v) for k, v in (obj.get("headers") or {}).items()}
        ra = headers.get("retry-after")
        retry_after = float(ra) if ra is not None and str(ra).strip().isdigit() else None
        if status in {408, 409, 425, 429, 500, 502, 503, 504}:
            return True, status, retry_after
        return False, status, retry_after
    except Exception:
        return False, None, None


def _parse_azure_http_error(e: Exception) -> dict | None:
    """Return our structured azure_openai_http_error dict if present."""
    try:
        obj = json.loads(str(e))
        if obj.get("type") == "azure_openai_http_error":
            return obj
    except Exception:
        return None
    return None


def _looks_like_max_tokens_unsupported(e: Exception) -> bool:
    s = str(e).lower()
    return "unsupported parameter" in s and "max_tokens" in s and "max_completion_tokens" in s


def _looks_like_temperature_unsupported(e: Exception) -> bool:
    s = str(e).lower()
    return "unsupported value" in s and "temperature" in s and ("default" in s or "only" in s)


# Knowledge placeholders when fields are not in the incident extract.
_LLM_JUDGE_KB_NA = "[Not available in incident extract]"

_LLM_JUDGE_SYSTEM = (
    "You are a neutral expert evaluator for IT service desk incident classification. "
    "Follow the user message rubric exactly. Populate the structured output: "
    "seven dimensions for WALLE and seven for DEX (each score is 0, 1, 2, 3, 4, or N); "
    "reasoning strings; D5/D6 example fields; per-model summary averages; final verdict. "
    "Do not claim one model is globally superior—this incident only."
)


def _build_llm_judge_incident_fields(
    row: pd.Series,
    *,
    safe_txt: Callable[[Any], str],
    columns: set[str],
) -> dict[str, str]:
    """Map merged CSV columns to the judge prompt incident placeholders."""

    def g(col: str) -> str:
        if col not in columns:
            return ""
        return safe_txt(row.get(col))

    def concat_labeled(pairs: list[tuple[str, str]]) -> str:
        lines = [f"{lab}: {txt}" for lab, txt in pairs if txt]
        return "\n".join(lines) if lines else "(none)"

    customer = g("INC_BRIEF_DESCRIPTION")
    if g("INC_UH_ESS_ERRORMSG"):
        customer = (customer + "\n" + f"INC_UH_ESS_ERRORMSG: {g('INC_UH_ESS_ERRORMSG')}").strip()

    long_desc = concat_labeled(
        [
            ("INC_ACTION", g("INC_ACTION")),
            ("INC_COMMENTS", g("INC_COMMENTS")),
            ("INC_UPDATE_ACTION", g("INC_UPDATE_ACTION")),
        ]
    )
    work_notes = concat_labeled(
        [
            ("INC_COMMENTS", g("INC_COMMENTS")),
            ("INC_UH_MONITORING_NOTES", g("INC_UH_MONITORING_NOTES")),
        ]
    )
    steps = concat_labeled(
        [
            ("INC_UPDATE_ACTION_ESS", g("INC_UPDATE_ACTION_ESS")),
            ("INC_UPDATE_ACTION", g("INC_UPDATE_ACTION")),
        ]
    )

    return {
        "inc_id": g("INC_ID") or "(unknown)",
        "inc_customer_states": customer or "(none)",
        "inc_long_description": long_desc or "(none)",
        "inc_resolution": g("INC_RESOLUTION") or "(none)",
        "inc_work_notes": work_notes or "(none)",
        "inc_steps_taken": steps or "(none)",
        "inc_kb_issue": _LLM_JUDGE_KB_NA,
        "inc_kb_fix": _LLM_JUDGE_KB_NA,
    }


def _llm_judge_user_prompt(*, incident: dict[str, str], walle: dict[str, str], dex: dict[str, str]) -> str:
    """Full rubric + incident + classifications (WALLE and DEX named explicitly)."""
    return f"""You are a neutral expert evaluator assessing the quality of two AI-generated
incident classification outputs — WALLE and DEX — for a single IT
service desk incident.

Your evaluation has two equally weighted objectives:

  OBJECTIVE 1 — ACCURACY
  Does the classification correctly reflect what happened in this incident,
  as evidenced by the incident record?

  OBJECTIVE 2 — BUSINESS ACTIONABILITY
  Does the classification produce outputs that a business can act on —
  specifically to drive automation, eliminate recurring incident causes,
  redesign processes, or reduce incident volume for that category?

A classification that is accurate but not actionable has limited business
value. A classification that is actionable but inaccurate will drive
automation in the wrong direction. Both objectives must be satisfied.

---

## INCIDENT RECORD

Incident ID:                  {incident["inc_id"]}
Customer stated:              {incident["inc_customer_states"]}
Long description:             {incident["inc_long_description"]}
Resolution applied:           {incident["inc_resolution"]}
Work notes:                   {incident["inc_work_notes"]}
Steps taken:                  {incident["inc_steps_taken"]}
Knowledge article — issue:    {incident["inc_kb_issue"]}
Knowledge article — fix:      {incident["inc_kb_fix"]}

---

## CLASSIFICATION OUTPUTS

WALLE:
  Domain (L1):      {walle["L1"]}
  Category (L2):    {walle["L2"]}
  Subcategory (L3): {walle["L3"]}
  Key Issue (L4):   {walle["L4"]}

DEX:
  Domain (L1):      {dex["L1"]}
  Category (L2):    {dex["L2"]}
  Subcategory (L3): {dex["L3"]}
  Key Issue (L4):   {dex["L4"]}

---

## EVALUATION INSTRUCTIONS

Score each model on the seven dimensions below.
Score each model INDEPENDENTLY — do not compare them to each other
when assigning scores. Both models can receive the same score on any
dimension.

SCORING SCALE:
  4 = Fully meets the dimension criteria
  3 = Mostly meets the criteria with minor gaps
  2 = Partially meets the criteria, notable gaps
  1 = Weakly meets the criteria, significant gaps
  0 = Does not meet the criteria or actively contradicts it
  N = Not applicable (field missing or dimension cannot be assessed)

GROUND RULES — read before scoring:
  - Evaluate ALL INC_ field groups with equal weight. Do not allow the
    customer statement alone to dominate your judgment. Resolution text,
    work notes, steps taken, and knowledge article references are equally
    valid evidence.
  - A null or missing L4 Key Issue field is NOT automatically a failure.
    If the incident narrative is ambiguous or sparse, null is a legitimate
    and calibrated response. Score it 2 unless the incident clearly had
    enough signal to classify, in which case score 1.
  - A populated L4 field is NOT automatically a success. If the label does
    not align with evidence or business use, score it low regardless of
    whether it is filled in.
  - Do not penalise a model for taxonomy vocabulary differences. Evaluate
    alignment to the incident and to business utility — not alignment to
    the other model's label format.

---

## SECTION 1 — ACCURACY DIMENSIONS (50% of total score)

DIMENSION 1 — Problem Identification Accuracy
Does the classification correctly identify what the actual problem was,
as evidenced by the customer statement and long description?
A good classification names the affected system, service, or component
and the nature of the failure the user experienced.

  WALLE score (0-4):
  Reasoning (2-3 sentences):

  DEX score (0-4):
  Reasoning (2-3 sentences):

---

DIMENSION 2 — Resolution Alignment
Does the classification align with how the incident was actually resolved,
as evidenced by the resolution text and work notes?
A good classification should not contradict the resolution path taken.
Consider: if you only read the classification, would you be pointed in the
right direction to fix this type of incident in future?

  WALLE score (0-4):
  Reasoning (2-3 sentences):

  DEX score (0-4):
  Reasoning (2-3 sentences):

---

DIMENSION 3 — Knowledge Article Consistency
Where a knowledge article is referenced, does the classification align
with the issue and resolution described in that article?
If no knowledge article is present, mark both models N.

  WALLE score (0-4 or N):
  Reasoning (2-3 sentences):

  DEX score (0-4 or N):
  Reasoning (2-3 sentences):

---

DIMENSION 4 — Taxonomy Internal Consistency
Are all four levels (L1 through L4) logically consistent with each other?
Does L2 follow naturally from L1? Does L3 follow from L2? Does L4 follow
from L3? A classification with internally contradictory levels will
produce unreliable aggregations and misleading trend reports regardless
of whether individual levels are accurate.

  WALLE score (0-4):
  Reasoning (2-3 sentences):

  DEX score (0-4):
  Reasoning (2-3 sentences):

---

## SECTION 2 — BUSINESS ACTIONABILITY DIMENSIONS (50% of total score)

For each dimension below, consider what a business operations team,
automation engineer, or service improvement manager could realistically
do with this classification output at scale.

---

DIMENSION 5 — Automation Potential
Could this classification directly trigger or inform an automated
resolution workflow, virtual agent script, or self-service deflection?

Consider:
  - Is the classified issue type specific enough to map to a known
    automated fix? (e.g. account unlock, VPN reset, cache clear,
    PIN reset, device compliance check)
  - Does the classification distinguish between incident types that
    require different automation paths, or does it group them into
    a category too broad to automate against?
  - If this classification were used as a routing label to an
    automation engine, would the engine know what to execute?

Penalise vague or catch-all labels only when the incident record
contained enough information to produce a specific, automatable label.
Do not penalise appropriate abstraction when the incident itself was
genuinely ambiguous.

  WALLE score (0-4):
  Reasoning (2-3 sentences):
  Example automation action this classification could enable (if any):

  DEX score (0-4):
  Reasoning (2-3 sentences):
  Example automation action this classification could enable (if any):

---

DIMENSION 6 — Process Improvement and Incident Reduction Signal
Could this classification, when aggregated across hundreds or thousands
of similar incidents, reveal a pattern that a process owner or IT
operations manager could act on to reduce incident volume?

Consider:
  - Does the classification reveal a root cause or failure mode that
    could be addressed upstream? (e.g. a recurring onboarding gap,
    a policy configuration that repeatedly locks users out, a software
    deployment that consistently fails on a specific device model)
  - Is the classification specific enough to distinguish actionable
    patterns from noise? A label of "Software / Error" tells a manager
    nothing. A label of "Windows Hello PIN Reset Required — post device
    replacement" tells them exactly where to intervene.
  - Does the classification capture enough context that a Problem
    Management team could use it to open a Problem ticket and drive
    a permanent fix?
  - Would grouping incidents by this classification produce meaningful
    cohorts for trend analysis, or would it lump unrelated incidents
    together?

  WALLE score (0-4):
  Reasoning (2-3 sentences):
  Example process improvement or incident reduction this could drive:

  DEX score (0-4):
  Reasoning (2-3 sentences):
  Example process improvement or incident reduction this could drive:

---

DIMENSION 7 — Executive Reporting and Strategic Decision Support
Could this classification contribute to meaningful executive-level
reporting on IT health, workforce productivity impact, and technology
investment priorities?

Consider:
  - Is the classification specific enough to appear as a meaningful
    category in a dashboard without requiring manual reclassification
    or post-processing?
  - Does it support trend analysis over time — i.e. could a leader
    track whether this incident type is increasing, decreasing, or
    shifting to different products?
  - Does it attribute the incident to a product, service, or platform
    in a way that could inform vendor management, contract decisions,
    or technology refresh priorities?
  - Would a non-technical executive understand what this classification
    represents, or would it require significant translation?

  WALLE score (0-4):
  Reasoning (2-3 sentences):

  DEX score (0-4):
  Reasoning (2-3 sentences):

---

## FINAL SCORING SUMMARY

List scores for all applicable dimensions and calculate averages.
Exclude any dimension marked N from both the numerator and denominator.
Show your working clearly in the structured output fields (applicable_dimensions lists which D1-D7 counted; averages are numeric).

WALLE and DEX each need:
  D1 Problem Identification through D7 Executive Reporting scores (or N),
  applicable_dimensions list,
  Section 1 average (D1-D4),
  Section 2 average (D5-D7),
  Overall average (equal weight over applicable dimensions).

---

## VERDICT

Declare a winner or a tie. A tie is declared if overall averages
are within 0.25 of each other.

State separately whether one model led on accuracy while the other
led on actionability — this split verdict is important because the
two objectives may be best served by different models in combination
rather than a single winner replacing the other.

Overall verdict:              [WALLE / DEX / Tie]
Accuracy leader:              [WALLE / DEX / Equal]
Actionability leader:         [WALLE / DEX / Equal]
Dimension with largest gap:   [D1-D7 or None]
One-sentence summary:

IMPORTANT — do not make any general claim about which model is
superior overall. Your verdict applies to this single incident only.
Cumulative conclusions across incidents must be drawn by a human
reviewer after all incidents are scored.

---

Your response must be provided ONLY through the structured output schema (no extra prose outside it).
"""


def _flatten_judge_out_for_excel(
    inc_id: str,
    walle_labels: dict[str, str],
    dex_labels: dict[str, str],
    obj: dict[str, Any],
) -> dict[str, Any]:
    """One wide row for metric_6_llm_judge_rows."""
    out: dict[str, Any] = {
        "INC_ID": inc_id,
        "WALLE_L1": walle_labels.get("L1"),
        "WALLE_L2": walle_labels.get("L2"),
        "WALLE_L3": walle_labels.get("L3"),
        "WALLE_L4": walle_labels.get("L4"),
        "DEX_L1": dex_labels.get("L1"),
        "DEX_L2": dex_labels.get("L2"),
        "DEX_L3": dex_labels.get("L3"),
        "DEX_L4": dex_labels.get("L4"),
    }
    w = obj.get("walle") or {}
    d = obj.get("dex") or {}
    dim_map = [
        ("d1_problem_identification", "D1"),
        ("d2_resolution_alignment", "D2"),
        ("d3_knowledge_article", "D3"),
        ("d4_taxonomy_consistency", "D4"),
        ("d5_automation", "D5"),
        ("d6_process_improvement", "D6"),
        ("d7_executive_reporting", "D7"),
    ]
    for field, tag in dim_map:
        wb = w.get(field) or {}
        db = d.get(field) or {}
        out[f"walle_{tag}_score"] = wb.get("score")
        out[f"walle_{tag}_reasoning"] = wb.get("reasoning")
        out[f"dex_{tag}_score"] = db.get("score")
        out[f"dex_{tag}_reasoning"] = db.get("reasoning")
        if tag == "D5":
            out["walle_D5_example_automation"] = wb.get("example_automation", "")
            out["dex_D5_example_automation"] = db.get("example_automation", "")
        if tag == "D6":
            out["walle_D6_example_process_improvement"] = wb.get("example_process_improvement", "")
            out["dex_D6_example_process_improvement"] = db.get("example_process_improvement", "")
    ws = obj.get("walle_summary") or {}
    ds = obj.get("dex_summary") or {}
    out["walle_applicable_dimensions"] = ws.get("applicable_dimensions")
    out["walle_section1_avg"] = ws.get("section1_average")
    out["walle_section2_avg"] = ws.get("section2_average")
    out["walle_overall_avg"] = ws.get("overall_average")
    out["dex_applicable_dimensions"] = ds.get("applicable_dimensions")
    out["dex_section1_avg"] = ds.get("section1_average")
    out["dex_section2_avg"] = ds.get("section2_average")
    out["dex_overall_avg"] = ds.get("overall_average")
    ver = obj.get("verdict") or {}
    out["verdict_overall"] = ver.get("overall_verdict")
    out["verdict_accuracy_leader"] = ver.get("accuracy_leader")
    out["verdict_actionability_leader"] = ver.get("actionability_leader")
    out["verdict_dimension_largest_gap"] = ver.get("dimension_largest_gap")
    out["verdict_one_sentence_summary"] = ver.get("one_sentence_summary")
    return out


def _create_llm_judge_agent() -> Any:
    """
    Build a single pydantic-ai Agent for WALLE vs DEX judge (same stack as L4).
    Reuse one agent across concurrent runs (asyncio.gather + Semaphore), like batch workers in BaseClassifier.
    """
    import os
    from pydantic import BaseModel, Field
    from pydantic_ai import Agent
    from pydantic_ai.models.openai import OpenAIChatModel, OpenAIChatModelSettings

    class DimScore(BaseModel):
        score: str = Field(..., description="Exactly one of: 0, 1, 2, 3, 4, N")
        reasoning: str = Field(..., description="2-3 sentences")

    class Dim5Score(BaseModel):
        score: str = Field(..., description="0-4 or N")
        reasoning: str = Field(..., description="2-3 sentences")
        example_automation: str = Field(default="", description="Short phrase or empty")

    class Dim6Score(BaseModel):
        score: str = Field(..., description="0-4 or N")
        reasoning: str = Field(..., description="2-3 sentences")
        example_process_improvement: str = Field(default="", description="Short phrase or empty")

    class ModelEvalScores(BaseModel):
        d1_problem_identification: DimScore
        d2_resolution_alignment: DimScore
        d3_knowledge_article: DimScore
        d4_taxonomy_consistency: DimScore
        d5_automation: Dim5Score
        d6_process_improvement: Dim6Score
        d7_executive_reporting: DimScore

    class ModelSummary(BaseModel):
        applicable_dimensions: str = Field(
            ...,
            description="Comma-separated e.g. D1,D2,D3,D4,D5,D6,D7 — exclude dimensions scored N",
        )
        section1_average: float = Field(..., description="Mean of D1-D4 numeric scores only, 0-4 scale")
        section2_average: float = Field(..., description="Mean of D5-D7 numeric scores only, 0-4 scale")
        overall_average: float = Field(
            ...,
            description="Equal-weight mean over all applicable numeric dimensions, 0-4 scale",
        )

    class VerdictOut(BaseModel):
        overall_verdict: str = Field(..., description="WALLE, DEX, or Tie")
        accuracy_leader: str = Field(..., description="WALLE, DEX, or Equal")
        actionability_leader: str = Field(..., description="WALLE, DEX, or Equal")
        dimension_largest_gap: str = Field(..., description="D1, D2, ... D7, or None")
        one_sentence_summary: str = Field(..., description="Single incident only")

    class JudgeOut(BaseModel):
        walle: ModelEvalScores
        dex: ModelEvalScores
        walle_summary: ModelSummary
        dex_summary: ModelSummary
        verdict: VerdictOut

    deployment = (os.getenv("AZURE_OPENAI_DEPLOYMENT_NAME") or os.getenv("AZURE_OPENAI_DEPLOYMENT") or "").strip()
    if not deployment:
        raise RuntimeError("Missing AZURE_OPENAI_DEPLOYMENT_NAME for pydantic-ai judge")

    if not (os.getenv("OPENAI_API_VERSION") or "").strip():
        v = (os.getenv("AZURE_OPENAI_API_VERSION") or "").strip()
        if v:
            os.environ["OPENAI_API_VERSION"] = v

    model = OpenAIChatModel(deployment, provider="azure")
    x_upstream_env = (os.getenv("X_UPSTREAM_ENV") or os.getenv("X-Upstream-Env") or "").strip()
    project_id = (os.getenv("PROJECT_ID") or os.getenv("projectId") or "").strip()
    extra_headers: dict[str, str] = {}
    if x_upstream_env:
        extra_headers.update(
            {
                "X-Upstream-Env": x_upstream_env,
                "projectId": project_id,
                "X-Model-Usage-Type": x_upstream_env,
                "modelUsageType": x_upstream_env,
            }
        )
    elif project_id:
        extra_headers["projectId"] = project_id

    settings = OpenAIChatModelSettings(extra_headers=extra_headers) if extra_headers else OpenAIChatModelSettings()
    return Agent(model, output_type=JudgeOut, model_settings=settings, system_prompt=_LLM_JUDGE_SYSTEM)


def _agent_run_result_to_judge_dict(res: Any) -> dict[str, Any]:
    if hasattr(res, "data"):
        out = res.data
    elif hasattr(res, "output"):
        out = res.output
    else:
        raise RuntimeError(f"Unexpected AgentRunResult shape: {type(res)}")
    if hasattr(out, "model_dump"):
        return out.model_dump()
    return dict(out)


async def _judge_one_incident_async(
    agent: Any,
    *,
    semaphore: Any,
    inc_id: str,
    user_prompt: str,
    wal: dict[str, str],
    dex: dict[str, str],
    max_retries: int,
    progress_lock: Any,
    progress_state: list[int],
    total: int,
) -> dict[str, Any]:
    """One incident under semaphore; retries with asyncio.sleep (L4-style transient handling)."""
    import asyncio

    async with semaphore:
        attempt = 0
        while True:
            try:
                res = await agent.run(user_prompt)
                obj = _agent_run_result_to_judge_dict(res)
                rec = _flatten_judge_out_for_excel(inc_id, wal, dex, obj)
                async with progress_lock:
                    progress_state[0] += 1
                    done = progress_state[0]
                    if done % max(1, total // 20) == 0 or done == total:
                        print(f"[LLM-judge] completed {done}/{total}", flush=True)
                return rec
            except Exception as e:
                attempt += 1
                retryable, status, retry_after = _is_retryable_azure_error(e)
                if retryable and attempt <= max_retries:
                    sleep_for = float(retry_after) if retry_after is not None else min(60.0, (2.0 ** min(attempt, 6)))
                    await asyncio.sleep(sleep_for)
                    continue
                async with progress_lock:
                    progress_state[0] += 1
                    done = progress_state[0]
                    if done % max(1, total // 20) == 0 or done == total:
                        print(f"[LLM-judge] completed {done}/{total}", flush=True)
                return {
                    "INC_ID": inc_id,
                    "error": str(e)[:800],
                    "status_code": status,
                    "attempts": attempt,
                }


async def _run_llm_judge_parallel(
    work_items: list[dict[str, Any]],
    *,
    workers: int,
    max_retries: int,
) -> list[dict[str, Any]]:
    """
    Parallel judge pass (same spirit as BaseClassifier._classify_single_worker):
    asyncio.Semaphore(workers) limits in-flight requests; asyncio.gather preserves input order.
    """
    import asyncio

    agent = _create_llm_judge_agent()
    workers = max(1, int(workers))
    semaphore = asyncio.Semaphore(workers)
    progress_lock = asyncio.Lock()
    progress_state = [0]
    total = len(work_items)

    tasks = [
        asyncio.create_task(
            _judge_one_incident_async(
                agent,
                semaphore=semaphore,
                inc_id=wi["inc_id"],
                user_prompt=wi["user_prompt"],
                wal=wi["wal"],
                dex=wi["dex"],
                max_retries=max_retries,
                progress_lock=progress_lock,
                progress_state=progress_state,
                total=total,
            )
        )
        for wi in work_items
    ]
    # Results are in the same order as work_items / tasks.
    return list(await asyncio.gather(*tasks))


async def _run_agent_prompts_parallel(
    *,
    agent: Any,
    prompts: list[str],
    workers: int,
    max_retries: int,
    progress_label: str,
) -> list[dict[str, Any]]:
    """
    Parallel runner for generic `agent.run(prompt)` calls.
    - Limits concurrency with asyncio.Semaphore(workers)
    - Preserves order via asyncio.gather
    - Retries 429/5xx using the same backoff logic as other LLM calls
    """
    import asyncio

    workers = max(1, int(workers))
    sem = asyncio.Semaphore(workers)
    lock = asyncio.Lock()
    state = [0]
    total = len(prompts)

    async def one(prompt: str) -> dict[str, Any]:
        async with sem:
            attempt = 0
            while True:
                try:
                    res = await agent.run(prompt)
                    obj = _agent_run_result_to_judge_dict(res)
                    async with lock:
                        state[0] += 1
                        done = state[0]
                        if total > 0 and (done % max(1, total // 20) == 0 or done == total):
                            print(f"[{progress_label}] completed {done}/{total}", flush=True)
                    return obj
                except Exception as e:
                    attempt += 1
                    retryable, status, retry_after = _is_retryable_azure_error(e)
                    if retryable and attempt <= max_retries:
                        backoff = min(60.0, (2.0 ** min(attempt, 6)))
                        sleep_for = float(retry_after) if retry_after is not None else float(backoff)
                        await asyncio.sleep(sleep_for)
                        continue
                    async with lock:
                        state[0] += 1
                        done = state[0]
                        if total > 0 and (done % max(1, total // 20) == 0 or done == total):
                            print(f"[{progress_label}] completed {done}/{total}", flush=True)
                    return {"error": str(e)[:800], "status_code": status, "attempts": attempt}

    tasks = [asyncio.create_task(one(p)) for p in prompts]
    return list(await asyncio.gather(*tasks))


def _prepend_sheet_explainability(
    xlsx_path: Path,
    per_sheet_lines: dict[str, list[str]],
) -> None:
    """
    Prepend human-readable explainability blocks to each sheet.

    We write the metric DataFrames with pandas first, then reopen the workbook and
    insert rows at the top of each sheet so the guidance travels with the tab.
    """
    wb = load_workbook(xlsx_path)
    for sheet_name, lines in per_sheet_lines.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        if not lines:
            continue

        # Insert guidance rows at top; keep a blank separator row.
        n_insert = len(lines) + 1
        ws.insert_rows(1, amount=n_insert)

        for i, line in enumerate(lines, start=1):
            ws.cell(row=i, column=1, value=line)

        # Make the first column readable for long lines.
        try:
            ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 120)
        except Exception:
            pass

        # Freeze panes at first data row (after explainability + header row).
        # Data header row will now begin at row n_insert + 1.
        ws.freeze_panes = ws.cell(row=n_insert + 2, column=1)

    wb.save(xlsx_path)


def main() -> int:
    p = argparse.ArgumentParser(description="Compute DEXter vs WALLE metrics from merged CSV.")
    p.add_argument("--in", dest="input_csv", required=True, help="Path to merged DEX+WALLE CSV")
    p.add_argument("--out", dest="output_xlsx", required=True, help="Path to output Excel report")
    p.add_argument(
        "--generic",
        default="unknown,other,unclassified,unclassified_l4,n/a,na,none",
        help="Comma-separated list of generic labels to treat as non-usable (default: common unknown buckets)",
    )
    p.add_argument(
        "--llm-judge",
        action="store_true",
        help="Enable LLM-as-judge metrics (requires Azure OpenAI env vars). Off by default.",
    )
    p.add_argument(
        "--llm-judge-n",
        type=int,
        default=0,
        help="How many rows to judge (0 means: if --llm-judge, judge up to 200).",
    )
    p.add_argument(
        "--llm-judge-seed",
        type=int,
        default=7,
        help="Random seed for judge row sampling (when limiting to llm-judge-n).",
    )
    p.add_argument(
        "--llm-judge-max-rpm",
        type=int,
        default=30,
        help=(
            "Legacy: was used for sequential spacing between judge calls. "
            "Parallel judge uses --llm-judge-workers for concurrency instead; this flag is currently unused."
        ),
    )
    p.add_argument(
        "--llm-judge-max-context-chars",
        type=int,
        default=6000,
        help="Max characters of incident text sent to judge (approx token budgeting).",
    )
    p.add_argument(
        "--llm-judge-max-retries",
        type=int,
        default=6,
        help="Max retries per judge call on 429/5xx with exponential backoff.",
    )
    p.add_argument(
        "--llm-judge-workers",
        type=int,
        default=5,
        help=(
            "Concurrent LLM judge tasks (asyncio.Semaphore), same idea as L4 batch workers. "
            "Default 5. Use 1 for sequential. Lower if you hit 429/rate limits."
        ),
    )
    args = p.parse_args()

    input_path = Path(args.input_csv)
    output_path = Path(args.output_xlsx)
    if not input_path.exists():
        raise FileNotFoundError(f"Input CSV not found: {input_path}")

    generic = {s.strip().lower() for s in str(args.generic).split(",") if s.strip()}

    df = pd.read_csv(input_path)

    # ---- Column contract normalization (readability) ----
    rename_map = {
        # single ID
        "DEX_INGEST_TICKET_ID": "INC_ID",
        "WALLE_IN_ID": "WALLE_IN_ID",

        # DEX hierarchy
        "DEX_CLASSIFICATION_DOMAIN": "DEX_L1",
        "DEX_CLASSIFICATION_CATEGORY": "DEX_L2",
        "DEX_CLASSIFICATION_SUBCATEGORY": "DEX_L3",
        "DEX_KEY_ISSUE_CATEGORY": "DEX_L4",

        # WALLE hierarchy
        "WALLE_AI_L1": "WALLE_L1",
        "WALLE_AI_L2": "WALLE_L2",
        "WALLE_AI_L3": "WALLE_L3",
        "WALLE_AI_L4": "WALLE_L4",

        # WALLE explainability fields (keep)
        "WALLE_VENDOR": "WALLE_VENDOR",
        "WALLE_AI_RATIONALE": "WALLE_AI_RATIONALE",
        "WALLE_AI_KEYWORDS": "WALLE_AI_KEYWORDS",
        "WALLE_AI_ROOT_CAUSE_INDICATOR": "WALLE_AI_ROOT_CAUSE_INDICATOR",
        "WALLE_AI_ROOT_CAUSE": "WALLE_AI_ROOT_CAUSE",
        "WALLE_AI_L4_CONFIDENCE": "WALLE_AI_L4_CONFIDENCE",
        "WALLE_AI_L4_RESOLUTION_ACTION": "WALLE_AI_L4_RESOLUTION_ACTION",
        "WALLE_AI_L4_ACTIONABLE": "WALLE_L4_ACTIONABLE",
        "WALLE_AI_L4_ACTIONABILITY_REASON": "WALLE_L4_ACTIONABILITY_REASON",
        "WALLE_AI_L4_RATIONALE": "WALLE_AI_L4_RATIONALE",

        # Incident context -> INC_ prefix (keep)
        "WALLE_BRIEF_DESCRIPTION": "INC_BRIEF_DESCRIPTION",
        "WALLE_ACTION": "INC_ACTION",
        "WALLE_RESOLUTION": "INC_RESOLUTION",
        "WALLE_OPENED_AT": "INC_OPENED_AT",
        "WALLE_CLOSED_AT": "INC_CLOSED_AT",
        "WALLE_UPDATE_ACTION_ESS": "INC_UPDATE_ACTION_ESS",
        "WALLE_UH_ESS_ERRORMSG": "INC_UH_ESS_ERRORMSG",
        "WALLE_UPDATE_ACTION": "INC_UPDATE_ACTION",
        "WALLE_COMMENTS": "INC_COMMENTS",
        "WALLE_UH_MONITORING_NOTES": "INC_UH_MONITORING_NOTES",
    }

    df = df.rename(columns={k: v for k, v in rename_map.items() if k in df.columns})

    # If both IDs exist, keep only INC_ID (but report mismatches).
    if "INC_ID" in df.columns and "WALLE_IN_ID" in df.columns:
        mism = (
            df["INC_ID"].astype(str).str.strip()
            != df["WALLE_IN_ID"].astype(str).str.strip()
        ).sum()
        print(f"[ID] mismatches INC_ID vs WALLE_IN_ID: {int(mism)}", flush=True)
        df = df.drop(columns=["WALLE_IN_ID"])

    # Keep only the columns we care about (contract).
    ordered_cols = [
        "INC_ID",
        "DEX_L1",
        "DEX_L2",
        "DEX_L3",
        "DEX_L4",
        "WALLE_L1",
        "WALLE_L2",
        "WALLE_L3",
        "WALLE_L4",
        "WALLE_VENDOR",
        "WALLE_AI_RATIONALE",
        "WALLE_AI_KEYWORDS",
        "WALLE_AI_ROOT_CAUSE_INDICATOR",
        "WALLE_AI_ROOT_CAUSE",
        "WALLE_AI_L4_CONFIDENCE",
        "WALLE_AI_L4_RESOLUTION_ACTION",
        "WALLE_L4_ACTIONABLE",
        "WALLE_L4_ACTIONABILITY_REASON",
        "WALLE_AI_L4_RATIONALE",
        "INC_BRIEF_DESCRIPTION",
        "INC_ACTION",
        "INC_RESOLUTION",
        "INC_OPENED_AT",
        "INC_CLOSED_AT",
        "INC_UPDATE_ACTION_ESS",
        "INC_UH_ESS_ERRORMSG",
        "INC_UPDATE_ACTION",
        "INC_COMMENTS",
        "INC_UH_MONITORING_NOTES",
    ]
    cols_present = [c for c in ordered_cols if c in df.columns]
    df_view = df[cols_present].copy()

    # Sanitize illegal control characters for Excel (openpyxl restriction).
    # These can appear in free-text fields and will crash the write.
    obj_cols = [c for c in df_view.columns if df_view[c].dtype == "object"]
    for c in obj_cols:
        df_view[c] = df_view[c].apply(
            lambda x: ILLEGAL_CHARACTERS_RE.sub("", x) if isinstance(x, str) else x
        )

    # ---- Shared setup for additional metrics (A–G): full 4-level paths ----
    df_view["walle_path"] = df_view.apply(
        lambda r: _build_path(r.get("WALLE_L1"), r.get("WALLE_L2"), r.get("WALLE_L3"), r.get("WALLE_L4"), generic),
        axis=1,
    )
    df_view["dex_path"] = df_view.apply(
        lambda r: _build_path(r.get("DEX_L1"), r.get("DEX_L2"), r.get("DEX_L3"), r.get("DEX_L4"), generic),
        axis=1,
    )

    # ---- Metric 1: coverage ----
    rows = len(df_view)
    out_rows: list[dict] = []
    for k in (1, 2, 3, 4):
        dex_col = f"DEX_L{k}"
        wal_col = f"WALLE_L{k}"
        dex_cov = None
        wal_cov = None
        if dex_col in df_view.columns:
            dex_cov = float(df_view[dex_col].apply(lambda v: _is_usable_label(v, generic)).mean() * 100)
        if wal_col in df_view.columns:
            wal_cov = float(df_view[wal_col].apply(lambda v: _is_usable_label(v, generic)).mean() * 100)
        out_rows.append(
            {
                "level": f"L{k}",
                "rows": rows,
                "dex_coverage_pct": dex_cov,
                "walle_coverage_pct": wal_cov,
            }
        )

    metric_1 = pd.DataFrame(out_rows)

    # ---- Metric 2: granularity / concentration / entropy ----
    # For each model + level, on usable labels only:
    # - unique label count
    # - top-10 share (%)
    # - entropy (higher means more spread; too low means overly concentrated)
    gran_rows: list[dict] = []
    top_labels_rows: list[dict] = []
    for model_prefix in ("DEX", "WALLE"):
        for k in (1, 2, 3, 4):
            col = f"{model_prefix}_L{k}"
            if col not in df_view.columns:
                continue
            usable = df_view.loc[_usable_mask(df_view[col], generic), col]
            norm = _normalize_label_series(usable)
            vc = norm.value_counts()
            unique_labels = int((vc > 0).sum())
            top10_share = _top_share(norm, 10)
            ent = _entropy_from_counts(vc)

            gran_rows.append(
                {
                    "model": model_prefix,
                    "level": f"L{k}",
                    "usable_rows": int(len(norm)),
                    "unique_labels": unique_labels,
                    "top10_share_pct": top10_share,
                    "entropy_bits": ent,
                }
            )

            # keep top 20 labels for explainability
            for label, cnt in vc.head(20).items():
                top_labels_rows.append(
                    {
                        "model": model_prefix,
                        "level": f"L{k}",
                        "label": label,
                        "count": int(cnt),
                        "pct_of_usable": float(cnt / len(norm) * 100.0) if len(norm) else 0.0,
                    }
                )

    metric_2 = pd.DataFrame(gran_rows)
    metric_2_top = pd.DataFrame(top_labels_rows)

    # ---- Metric 3: agreement + disagreements (DEX vs WALLE) ----
    # Exact match at each level, on rows where BOTH sides have usable labels.
    agree_rows: list[dict] = []
    disagree_rows: list[dict] = []
    for k in (1, 2, 3, 4):
        dex_col = f"DEX_L{k}"
        wal_col = f"WALLE_L{k}"
        if dex_col not in df_view.columns or wal_col not in df_view.columns:
            continue
        dex_usable = _usable_mask(df_view[dex_col], generic)
        wal_usable = _usable_mask(df_view[wal_col], generic)
        both = dex_usable & wal_usable
        both_n = int(both.sum())
        if both_n == 0:
            agree_rows.append(
                {
                    "level": f"L{k}",
                    "rows_total": rows,
                    "rows_both_usable": 0,
                    "exact_match_pct": None,
                }
            )
            continue

        dex_norm = _normalize_label_series(df_view.loc[both, dex_col])
        wal_norm = _normalize_label_series(df_view.loc[both, wal_col])
        match = (dex_norm == wal_norm)
        match_pct = float(match.mean() * 100.0)

        agree_rows.append(
            {
                "level": f"L{k}",
                "rows_total": rows,
                "rows_both_usable": both_n,
                "exact_match_pct": match_pct,
            }
        )

        # Top disagreement pairs
        pairs = pd.DataFrame({"dex": dex_norm, "walle": wal_norm})
        pairs = pairs[pairs["dex"] != pairs["walle"]]
        if not pairs.empty:
            pair_counts = pairs.value_counts().head(25)
            for (dex_label, wal_label), cnt in pair_counts.items():
                disagree_rows.append(
                    {
                        "level": f"L{k}",
                        "dex_label": dex_label,
                        "walle_label": wal_label,
                        "count": int(cnt),
                        "pct_of_both_usable": float(cnt / both_n * 100.0),
                    }
                )

    metric_3 = pd.DataFrame(agree_rows)
    metric_3_disagree = pd.DataFrame(disagree_rows)

    # ---- Metric 4: WALLE explainability / actionability quality (no DEX equivalent) ----
    # Goal: quantify whether "actionability + rationale/confidence/root-cause fields" are present and consistent.
    m4_rows: list[dict] = []
    m4_conf_bins_rows: list[dict] = []
    m4_cols = {
        "WALLE_L4": "walle_l4",
        "WALLE_L4_ACTIONABLE": "actionable",
        "WALLE_L4_ACTIONABILITY_REASON": "actionability_reason",
        "WALLE_AI_L4_CONFIDENCE": "l4_confidence",
        "WALLE_AI_L4_RATIONALE": "l4_rationale",
        "WALLE_AI_RATIONALE": "rationale",
        "WALLE_AI_KEYWORDS": "keywords",
        "WALLE_AI_ROOT_CAUSE_INDICATOR": "root_cause_indicator",
        "WALLE_AI_ROOT_CAUSE": "root_cause",
        "WALLE_AI_L4_RESOLUTION_ACTION": "l4_resolution_action",
    }
    m4_present = {k: v for k, v in m4_cols.items() if k in df_view.columns}
    if m4_present:
        w = df_view.rename(columns=m4_present).copy()
        w["l4_usable"] = _usable_mask(w["walle_l4"], generic) if "walle_l4" in w.columns else False
        w["actionable_bool"] = w["actionable"].apply(_is_trueish) if "actionable" in w.columns else False
        if "l4_confidence" in w.columns:
            w["l4_confidence_num"] = pd.to_numeric(w["l4_confidence"], errors="coerce")

        # Define presence flags
        def _nonempty(col: str) -> pd.Series:
            if col not in w.columns:
                return pd.Series([False] * len(w))
            return w[col].apply(lambda v: (v is not None) and (not pd.isna(v)) and (str(v).strip() != ""))

        flags = {
            "has_rationale": _nonempty("rationale"),
            "has_l4_rationale": _nonempty("l4_rationale"),
            "has_keywords": _nonempty("keywords"),
            "has_root_cause_indicator": _nonempty("root_cause_indicator"),
            "has_root_cause": _nonempty("root_cause"),
            "has_l4_resolution_action": _nonempty("l4_resolution_action"),
            "has_actionability_reason": _nonempty("actionability_reason"),
            "has_l4_confidence": _nonempty("l4_confidence"),
            "has_l4_confidence_num": (~w.get("l4_confidence_num", pd.Series([pd.NA] * len(w))).isna())
            if "l4_confidence_num" in w.columns
            else pd.Series([False] * len(w)),
        }

        # Overall and conditional rates (only where L4 is usable)
        base = w["l4_usable"] if "l4_usable" in w.columns else pd.Series([True] * len(w))
        actionable = w["actionable_bool"] if "actionable_bool" in w.columns else pd.Series([False] * len(w))
        not_actionable = (~actionable)

        def _add_slice(slice_name: str, mask: pd.Series) -> None:
            denom = int(mask.sum())
            m4_rows.append({"slice": slice_name, "rows": denom, "note": "All % are within this slice"})
            if denom == 0:
                return
            for fname, fmask in flags.items():
                m4_rows.append(
                    {
                        "slice": slice_name,
                        "rows": denom,
                        "field": fname,
                        "pct_present": _pct(fmask[mask]),
                    }
                )
            if "l4_confidence_num" in w.columns:
                conf = w.loc[mask, "l4_confidence_num"]
                m4_rows.append(
                    {
                        "slice": slice_name,
                        "rows": denom,
                        "field": "l4_confidence_mean",
                        "pct_present": None,
                        "value": float(conf.mean()) if conf.notna().any() else None,
                    }
                )
                m4_rows.append(
                    {
                        "slice": slice_name,
                        "rows": denom,
                        "field": "l4_confidence_median",
                        "pct_present": None,
                        "value": float(conf.median()) if conf.notna().any() else None,
                    }
                )

        _add_slice("all_rows", pd.Series([True] * len(w)))
        _add_slice("walle_l4_usable", base)
        _add_slice("walle_l4_usable__actionable_true", base & actionable)
        _add_slice("walle_l4_usable__actionable_false", base & not_actionable)

        # Confidence histogram bins (if numeric)
        if "l4_confidence_num" in w.columns:
            conf_all = w.loc[base, "l4_confidence_num"].dropna()
            if not conf_all.empty:
                bins = [-float("inf"), 0, 0.25, 0.5, 0.75, 1.0, float("inf")]
                labels = ["<0", "0-0.25", "0.25-0.5", "0.5-0.75", "0.75-1.0", ">1.0"]
                b = pd.cut(conf_all, bins=bins, labels=labels, include_lowest=True)
                vc = b.value_counts().reindex(labels, fill_value=0)
                total = int(vc.sum())
                for lab, cnt in vc.items():
                    m4_conf_bins_rows.append(
                        {"slice": "walle_l4_usable", "confidence_bin": str(lab), "count": int(cnt), "pct": float(cnt / total * 100.0)}
                    )

    metric_4 = pd.DataFrame(m4_rows)
    metric_4_conf = pd.DataFrame(m4_conf_bins_rows)

    # ---- Metric 5: Stratified agreement/disagreement by WALLE strata and month buckets ----
    # Goal: find where the disagreement concentrates (e.g., certain WALLE_L1/L2/L3 or certain months).
    m5_rows: list[dict] = []
    m5_pair_rows: list[dict] = []
    if "INC_CLOSED_AT" in df_view.columns:
        month_bucket = _to_month_bucket(df_view["INC_CLOSED_AT"])
    elif "INC_OPENED_AT" in df_view.columns:
        month_bucket = _to_month_bucket(df_view["INC_OPENED_AT"])
    else:
        month_bucket = pd.Series([""] * len(df_view))
    df_m5 = df_view.copy()
    df_m5["month_bucket"] = month_bucket

    # Define strata keys (use what exists)
    strata_cols = [c for c in ["WALLE_L1", "WALLE_L2", "WALLE_L3"] if c in df_m5.columns]
    # If no strata columns exist, we can’t localize; keep empty metric_5.
    if strata_cols:
        # Use only rows where WALLE L1/L2/L3 are usable (as requested previously for sampling).
        strata_mask = pd.Series([True] * len(df_m5))
        for c in strata_cols:
            strata_mask &= _usable_mask(df_m5[c], generic)

        # Stratify agreement at each level by (WALLE_L1, month_bucket) and by full (WALLE_L1,L2,L3) if present.
        # Keep only groups with enough mutually-usable rows to be meaningful.
        min_both_usable = 20

        def _compute_group_agreement(group_df: pd.DataFrame, group_key: dict) -> None:
            for k in (1, 2, 3, 4):
                dex_col = f"DEX_L{k}"
                wal_col = f"WALLE_L{k}"
                if dex_col not in group_df.columns or wal_col not in group_df.columns:
                    continue
                dex_ok = _usable_mask(group_df[dex_col], generic)
                wal_ok = _usable_mask(group_df[wal_col], generic)
                both = dex_ok & wal_ok
                both_n = int(both.sum())
                if both_n < min_both_usable:
                    continue
                dex_norm = _normalize_label_series(group_df.loc[both, dex_col])
                wal_norm = _normalize_label_series(group_df.loc[both, wal_col])
                match_pct = float((dex_norm == wal_norm).mean() * 100.0)
                m5_rows.append(
                    {
                        **group_key,
                        "level": f"L{k}",
                        "rows_group": int(len(group_df)),
                        "rows_both_usable": both_n,
                        "exact_match_pct": match_pct,
                    }
                )

                # Top disagreement pair inside this group (to quickly explain “why low”)
                pairs = pd.DataFrame({"dex": dex_norm, "walle": wal_norm})
                pairs = pairs[pairs["dex"] != pairs["walle"]]
                if not pairs.empty:
                    top = pairs.value_counts().head(1)
                    for (dex_label, wal_label), cnt in top.items():
                        m5_pair_rows.append(
                            {
                                **group_key,
                                "level": f"L{k}",
                                "dex_label": dex_label,
                                "walle_label": wal_label,
                                "count": int(cnt),
                                "pct_of_both_usable": float(cnt / both_n * 100.0),
                            }
                        )

        # Grouping 1: by WALLE_L1 + month_bucket (if month available)
        group_cols_1 = [strata_cols[0], "month_bucket"]
        for (l1, mb), g in df_m5.loc[strata_mask].groupby(group_cols_1, dropna=False):
            if str(mb).strip() == "":
                continue
            _compute_group_agreement(g, {"stratum": "WALLE_L1 x month", "WALLE_L1": l1, "month_bucket": mb})

        # Grouping 2: by WALLE_L1/L2/L3 combo (no month)
        group_cols_2 = strata_cols.copy()
        for keys, g in df_m5.loc[strata_mask].groupby(group_cols_2, dropna=False):
            key_dict = {"stratum": "WALLE_L1/L2/L3"} if len(group_cols_2) >= 2 else {"stratum": "WALLE_L1"}
            if not isinstance(keys, tuple):
                keys = (keys,)
            for col, val in zip(group_cols_2, keys):
                key_dict[col] = val
            _compute_group_agreement(g, key_dict)

    metric_5 = pd.DataFrame(m5_rows)
    metric_5_pairs = pd.DataFrame(m5_pair_rows)

    # ---- Metric 6: LLM-as-judge (optional) ----
    # Full rubric: WALLE vs DEX on 7 dimensions + summaries + verdict (see _llm_judge_user_prompt).
    metric_6_rows = pd.DataFrame([])
    metric_6_summary = pd.DataFrame([])
    llm_lines: list[str] = []
    if bool(args.llm_judge):
        import os

        endpoint = os.getenv("AZURE_OPENAI_ENDPOINT", "").strip()
        api_key = os.getenv("AZURE_OPENAI_API_KEY", "").strip()
        deployment = os.getenv("AZURE_OPENAI_DEPLOYMENT_NAME", os.getenv("AZURE_OPENAI_DEPLOYMENT", "")).strip()

        if not (endpoint and api_key and deployment):
            reason = "missing one of AZURE_OPENAI_ENDPOINT / AZURE_OPENAI_API_KEY / AZURE_OPENAI_DEPLOYMENT_NAME"
            llm_lines.append(f"- LLM judge skipped: {reason}.")
            # Make the sheets non-empty so it's obvious in the workbook itself.
            metric_6_rows = pd.DataFrame(
                [
                    {
                        "status": "skipped",
                        "reason": reason,
                        "AZURE_OPENAI_ENDPOINT_set": bool(endpoint),
                        "AZURE_OPENAI_API_KEY_set": bool(api_key),
                        "AZURE_OPENAI_DEPLOYMENT_NAME_set": bool(deployment),
                    }
                ]
            )
            metric_6_summary = pd.DataFrame([{"status": "skipped", "reason": reason}])
        else:
            n_default = 200
            judge_n = int(args.llm_judge_n) if int(args.llm_judge_n) > 0 else n_default
            judge_n = min(judge_n, len(df_view))

            # Prefer judging rows where both DEX_L4 and WALLE_L4 are usable, since that's most actionable.
            mask_both_l4 = pd.Series([True] * len(df_view))
            if "DEX_L4" in df_view.columns:
                mask_both_l4 &= _usable_mask(df_view["DEX_L4"], generic)
            if "WALLE_L4" in df_view.columns:
                mask_both_l4 &= _usable_mask(df_view["WALLE_L4"], generic)
            candidates = df_view[mask_both_l4].copy()
            if candidates.empty:
                candidates = df_view.copy()

            if len(candidates) > judge_n:
                candidates = candidates.sample(n=judge_n, random_state=int(args.llm_judge_seed))

            def _safe_txt(v) -> str:
                if v is None or pd.isna(v):
                    return ""
                s = str(v)
                s = ILLEGAL_CHARACTERS_RE.sub("", s)
                return s.strip()

            colset = set(candidates.columns)
            work_items: list[dict[str, Any]] = []
            for _idx, row in candidates.iterrows():
                inc_id = _safe_txt(row.get("INC_ID"))
                incident_fields = _build_llm_judge_incident_fields(row, safe_txt=_safe_txt, columns=colset)
                dex = {
                    "L1": _safe_txt(row.get("DEX_L1")),
                    "L2": _safe_txt(row.get("DEX_L2")),
                    "L3": _safe_txt(row.get("DEX_L3")),
                    "L4": _safe_txt(row.get("DEX_L4")),
                }
                wal = {
                    "L1": _safe_txt(row.get("WALLE_L1")),
                    "L2": _safe_txt(row.get("WALLE_L2")),
                    "L3": _safe_txt(row.get("WALLE_L3")),
                    "L4": _safe_txt(row.get("WALLE_L4")),
                }
                user = _llm_judge_user_prompt(incident=incident_fields, walle=wal, dex=dex)
                max_ctx = int(args.llm_judge_max_context_chars)
                if max_ctx > 0 and len(user) > max_ctx:
                    user = user[:max_ctx] + "\n\n[TRUNCATED: user prompt exceeded --llm-judge-max-context-chars]"
                work_items.append({"inc_id": inc_id, "user_prompt": user, "wal": wal, "dex": dex})

            import asyncio

            workers = max(1, int(args.llm_judge_workers))
            if not work_items:
                judge_out = []
            else:
                print(
                    f"[LLM-judge] starting {len(work_items)} incidents with workers={workers} "
                    f"(order preserved via asyncio.gather)",
                    flush=True,
                )
                judge_out = asyncio.run(
                    _run_llm_judge_parallel(
                        work_items,
                        workers=workers,
                        max_retries=int(args.llm_judge_max_retries),
                    )
                )
            failures = sum(1 for r in judge_out if r.get("error"))

            metric_6_rows = pd.DataFrame(judge_out)
            total_judged = int(len(metric_6_rows))
            llm_lines.append(
                f"- Judged rows: {total_judged} (failures: {failures}); parallel workers={workers}."
            )

            # Aggregate verdicts and mean reported averages (successful rows only).
            summary_rows: list[dict] = []
            ok = metric_6_rows
            if "error" in ok.columns:
                ok = ok[ok["error"].isna()]
            if "verdict_overall" in ok.columns:
                ok = ok[ok["verdict_overall"].notna()]

            n_ok = int(len(ok))
            if n_ok > 0:

                def _norm_v(s: str) -> str:
                    t = str(s).strip().upper()
                    if t in ("TIE",):
                        return "Tie"
                    if "WALLE" in t or t == "WALL-E":
                        return "WALLE"
                    if "DEX" in t:
                        return "DEX"
                    return t

                vo = ok["verdict_overall"].map(_norm_v)
                summary_rows.append(
                    {
                        "metric": "verdict_overall",
                        "rows": n_ok,
                        "WALLE_pct": float((vo == "WALLE").sum() / n_ok * 100.0),
                        "DEX_pct": float((vo == "DEX").sum() / n_ok * 100.0),
                        "Tie_pct": float((vo == "Tie").sum() / n_ok * 100.0),
                    }
                )
                if "verdict_accuracy_leader" in ok.columns:
                    va = ok["verdict_accuracy_leader"].map(_norm_v)
                    summary_rows.append(
                        {
                            "metric": "verdict_accuracy_leader",
                            "rows": n_ok,
                            "WALLE_pct": float((va == "WALLE").sum() / n_ok * 100.0),
                            "DEX_pct": float((va == "DEX").sum() / n_ok * 100.0),
                            "Equal_pct": float((va == "EQUAL").sum() / n_ok * 100.0),
                        }
                    )
                if "verdict_actionability_leader" in ok.columns:
                    vb = ok["verdict_actionability_leader"].map(_norm_v)
                    summary_rows.append(
                        {
                            "metric": "verdict_actionability_leader",
                            "rows": n_ok,
                            "WALLE_pct": float((vb == "WALLE").sum() / n_ok * 100.0),
                            "DEX_pct": float((vb == "DEX").sum() / n_ok * 100.0),
                            "Equal_pct": float((vb == "EQUAL").sum() / n_ok * 100.0),
                        }
                    )
                for col_avg, label in [
                    ("walle_overall_avg", "mean_walle_overall_avg"),
                    ("dex_overall_avg", "mean_dex_overall_avg"),
                    ("walle_section1_avg", "mean_walle_section1_avg"),
                    ("dex_section1_avg", "mean_dex_section1_avg"),
                    ("walle_section2_avg", "mean_walle_section2_avg"),
                    ("dex_section2_avg", "mean_dex_section2_avg"),
                ]:
                    if col_avg in ok.columns:
                        s = pd.to_numeric(ok[col_avg], errors="coerce")
                        if s.notna().any():
                            summary_rows.append({"metric": label, "rows": int(s.notna().sum()), "mean": float(s.mean())})
                if "walle_overall_avg" in ok.columns and "dex_overall_avg" in ok.columns:
                    w = pd.to_numeric(ok["walle_overall_avg"], errors="coerce")
                    d = pd.to_numeric(ok["dex_overall_avg"], errors="coerce")
                    m = (w - d).dropna()
                    if not m.empty:
                        summary_rows.append(
                            {
                                "metric": "mean_walle_minus_dex_overall_avg",
                                "rows": int(len(m)),
                                "mean": float(m.mean()),
                            }
                        )
                if "verdict_dimension_largest_gap" in ok.columns:
                    vc = ok["verdict_dimension_largest_gap"].astype(str).value_counts().head(5)
                    for gap, cnt in vc.items():
                        summary_rows.append(
                            {
                                "metric": "verdict_dimension_largest_gap_top",
                                "dimension": gap,
                                "count": int(cnt),
                                "pct_of_rows": float(cnt / n_ok * 100.0),
                            }
                        )
            metric_6_summary = pd.DataFrame(summary_rows)
    else:
        # If the flag wasn't enabled, still leave a visible marker row in the sheets.
        metric_6_rows = pd.DataFrame([{"status": "disabled", "reason": "Run with --llm-judge (workflow input llm_judge=true)"}])
        metric_6_summary = pd.DataFrame([{"status": "disabled", "reason": "Run with --llm-judge (workflow input llm_judge=true)"}])

    # ---- Per-sheet explainability blocks (inserted into each tab) ----
    usable_def = (
        "Usable label = non-NULL, non-blank, and not in generic buckets: "
        f"{', '.join(sorted(generic))}."
    )

    # Coverage deltas (computed from this run)
    cov_lines: list[str] = []
    if not metric_1.empty:
        for r in metric_1.itertuples(index=False):
            if r.dex_coverage_pct is None or r.walle_coverage_pct is None:
                continue
            delta = float(r.walle_coverage_pct) - float(r.dex_coverage_pct)
            cov_lines.append(f"- {r.level}: DEX={float(r.dex_coverage_pct):.1f}% | WALLE={float(r.walle_coverage_pct):.1f}% | WALLE-DEX={delta:+.1f} pp")

    # Granularity comparisons per level
    gran_lines: list[str] = []
    if not metric_2.empty:
        for lvl in ("L1", "L2", "L3", "L4"):
            d = metric_2[(metric_2["model"] == "DEX") & (metric_2["level"] == lvl)]
            w = metric_2[(metric_2["model"] == "WALLE") & (metric_2["level"] == lvl)]
            if d.empty or w.empty:
                continue
            d = d.iloc[0]
            w = w.iloc[0]
            gran_lines.append(
                f"- {lvl}: top10_share DEX={float(d['top10_share_pct']):.1f}% vs WALLE={float(w['top10_share_pct']):.1f}% | "
                f"entropy DEX={float(d['entropy_bits']):.2f} vs WALLE={float(w['entropy_bits']):.2f} | "
                f"unique_labels DEX={int(d['unique_labels'])} vs WALLE={int(w['unique_labels'])}"
            )

    # Agreement lines per level
    agree_lines: list[str] = []
    if not metric_3.empty:
        for r in metric_3.itertuples(index=False):
            if r.exact_match_pct is None:
                agree_lines.append(f"- {r.level}: rows_both_usable=0 (agreement not computed)")
            else:
                agree_lines.append(
                    f"- {r.level}: exact_match={float(r.exact_match_pct):.1f}% on rows_both_usable={int(r.rows_both_usable)} (of rows_total={int(r.rows_total)})"
                )

    # Metric 4 inference lines (computed)
    m4_lines: list[str] = []
    if not metric_4.empty:
        try:
            base = metric_4[(metric_4["slice"] == "walle_l4_usable") & (metric_4["field"] == "has_l4_rationale")]
            act = metric_4[(metric_4["slice"] == "walle_l4_usable__actionable_true") & (metric_4["field"] == "has_actionability_reason")]
            conf = metric_4[(metric_4["slice"] == "walle_l4_usable") & (metric_4["field"] == "l4_confidence_mean")]
            if not base.empty and "pct_present" in base.columns:
                m4_lines.append(f"- WALLE_L4 usable: % with L4 rationale present = {float(base['pct_present'].iloc[0]):.1f}%")
            if not act.empty and "pct_present" in act.columns:
                m4_lines.append(f"- Actionable=True: % with actionability reason present = {float(act['pct_present'].iloc[0]):.1f}%")
            if not conf.empty and "value" in conf.columns and pd.notna(conf["value"].iloc[0]):
                m4_lines.append(f"- WALLE_L4 usable: mean L4 confidence = {float(conf['value'].iloc[0]):.3f}")
        except Exception:
            pass

    # Metric 5 inference lines (computed)
    m5_lines: list[str] = []
    if not metric_5.empty:
        # Show worst 5 strata for L4 (lowest agreement) to focus investigation.
        try:
            l4 = metric_5[metric_5["level"] == "L4"].copy()
            if not l4.empty:
                l4 = l4.sort_values(["exact_match_pct", "rows_both_usable"], ascending=[True, False]).head(5)
                for r in l4.itertuples(index=False):
                    parts = []
                    for c in ["stratum", "WALLE_L1", "WALLE_L2", "WALLE_L3", "month_bucket"]:
                        if hasattr(r, c) and getattr(r, c) is not None and str(getattr(r, c)).strip() != "" and c in l4.columns:
                            parts.append(f"{c}={getattr(r, c)}")
                    m5_lines.append(
                        f"- L4 low-agreement stratum: {', '.join(parts)} | match={float(r.exact_match_pct):.1f}% on both_usable={int(r.rows_both_usable)}"
                    )
        except Exception:
            pass

    # =====================================================================
    # Additional metrics (A–G) on full 4-level paths (walle_path/dex_path)
    # Run order: C → D → F → E → A → B → G
    # =====================================================================

    blocks_to_append: dict[str, list[dict]] = {}

    # ---- Metric C: MBO clusters ----
    CATCH_ALL_L3 = [
        "Error / Failure",
        "Other",
        "Unknown",
        "Error",
        "General",
        "Application/Service",
        "Desktop Related",
        "Other RFI",
    ]

    def _metric_c_for(model: str) -> pd.DataFrame:
        if model == "WALLE":
            path_col = "walle_path"
            l1, l2, l3, l4 = "WALLE_L1", "WALLE_L2", "WALLE_L3", "WALLE_L4"
        else:
            path_col = "dex_path"
            l1, l2, l3, l4 = "DEX_L1", "DEX_L2", "DEX_L3", "DEX_L4"
        base = df_view[[path_col, l1, l2, l3, l4]].copy()
        base = base[base[path_col].notna()]
        total = int(len(base))
        if total == 0:
            return pd.DataFrame([])
        g = (
            base.groupby([path_col, l1, l2, l3, l4], dropna=False)
            .size()
            .reset_index(name="incident_count")
            .sort_values("incident_count", ascending=False)
            .reset_index(drop=True)
        )
        g["model"] = model
        g["rank_by_volume"] = (g.index + 1).astype(int)
        g["full_path"] = g[path_col]
        g["l1_domain"] = g[l1]
        g["l2_category"] = g[l2]
        g["l3_subcategory"] = g[l3]
        g["l4_key_issue"] = g[l4]
        g["pct_of_total"] = g["incident_count"] / float(total) * 100.0
        g["cumulative_pct"] = g["pct_of_total"].cumsum()
        g["catch_all_flag"] = g["l3_subcategory"].astype(str).isin(CATCH_ALL_L3)
        g["mbo_viable"] = (g["incident_count"] >= 3) & (~g["catch_all_flag"]) & g["l4_key_issue"].apply(lambda v: _is_usable_label(v, generic))
        return g[
            [
                "model",
                "rank_by_volume",
                "full_path",
                "l1_domain",
                "l2_category",
                "l3_subcategory",
                "l4_key_issue",
                "incident_count",
                "pct_of_total",
                "cumulative_pct",
                "catch_all_flag",
                "mbo_viable",
            ]
        ]

    metric_C = pd.concat([_metric_c_for("WALLE"), _metric_c_for("DEX")], ignore_index=True)

    def _pareto_summary(df_model: pd.DataFrame) -> dict[str, Any]:
        if df_model.empty:
            return {
                "total_incidents": 0,
                "unique_paths": 0,
                "paths_to_cover_50pct": None,
                "paths_to_cover_80pct": None,
                "pct_incidents_in_viable_clusters": None,
                "pct_incidents_in_catch_all": None,
                "largest_cluster_path": None,
                "largest_cluster_pct": None,
            }
        total_incidents = int(df_model["incident_count"].sum())
        unique_paths = int(df_model["full_path"].nunique())
        paths_50 = int((df_model["cumulative_pct"] <= 50.0).sum()) + 1
        paths_80 = int((df_model["cumulative_pct"] <= 80.0).sum()) + 1
        pct_viable = float(df_model.loc[df_model["mbo_viable"], "incident_count"].sum() / total_incidents * 100.0) if total_incidents else None
        pct_catch = float(df_model.loc[df_model["catch_all_flag"], "incident_count"].sum() / total_incidents * 100.0) if total_incidents else None
        largest = df_model.head(1)
        return {
            "total_incidents": total_incidents,
            "unique_paths": unique_paths,
            "paths_to_cover_50pct": paths_50,
            "paths_to_cover_80pct": paths_80,
            "pct_incidents_in_viable_clusters": pct_viable,
            "pct_incidents_in_catch_all": pct_catch,
            "largest_cluster_path": str(largest["full_path"].iloc[0]) if not largest.empty else None,
            "largest_cluster_pct": float(largest["pct_of_total"].iloc[0]) if not largest.empty else None,
        }

    if not metric_C.empty:
        wsum = _pareto_summary(metric_C[metric_C["model"] == "WALLE"])
        dsum = _pareto_summary(metric_C[metric_C["model"] == "DEX"])
        blocks_to_append["metric_C_mbo_clusters"] = [
            {
                "title": "Pareto summary (label | WALLE value | DEX value)",
                "headers": ["label", "WALLE", "DEX"],
                "rows": [
                    ["total_incidents", wsum["total_incidents"], dsum["total_incidents"]],
                    ["unique_paths", wsum["unique_paths"], dsum["unique_paths"]],
                    ["paths_to_cover_50pct", wsum["paths_to_cover_50pct"], dsum["paths_to_cover_50pct"]],
                    ["paths_to_cover_80pct", wsum["paths_to_cover_80pct"], dsum["paths_to_cover_80pct"]],
                    ["pct_incidents_in_viable_clusters", wsum["pct_incidents_in_viable_clusters"], dsum["pct_incidents_in_viable_clusters"]],
                    ["pct_incidents_in_catch_all", wsum["pct_incidents_in_catch_all"], dsum["pct_incidents_in_catch_all"]],
                    ["largest_cluster_path", wsum["largest_cluster_path"], dsum["largest_cluster_path"]],
                    ["largest_cluster_pct", wsum["largest_cluster_pct"], dsum["largest_cluster_pct"]],
                ],
            }
        ]

    # ---- Metric D: path stability (TF-IDF similar pairs) ----
    try:
        from sklearn.feature_extraction.text import TfidfVectorizer
        from sklearn.metrics.pairwise import cosine_similarity

        texts = df_view.get("INC_BRIEF_DESCRIPTION", pd.Series([""] * len(df_view))).fillna("").astype(str).tolist()
        tfidf = TfidfVectorizer(max_features=300, stop_words="english")
        mat = tfidf.fit_transform(texts)
        sim = cosine_similarity(mat)
        THRESHOLD = 0.35
        d_rows: list[dict] = []
        n = len(df_view)
        for i in range(n):
            for j in range(i + 1, n):
                s = float(sim[i, j])
                if s >= THRESHOLD:
                    d_rows.append(
                        {
                            "inc_id_a": df_view["INC_ID"].iloc[i],
                            "inc_id_b": df_view["INC_ID"].iloc[j],
                            "text_similarity": s,
                            "walle_path_a": df_view["walle_path"].iloc[i],
                            "walle_path_b": df_view["walle_path"].iloc[j],
                            "walle_path_match": str(df_view["walle_path"].iloc[i]) == str(df_view["walle_path"].iloc[j]),
                            "walle_l1_match": str(df_view.get("WALLE_L1", pd.Series([""])).iloc[i]) == str(df_view.get("WALLE_L1", pd.Series([""])).iloc[j]),
                            "dex_path_a": df_view["dex_path"].iloc[i],
                            "dex_path_b": df_view["dex_path"].iloc[j],
                            "dex_path_match": str(df_view["dex_path"].iloc[i]) == str(df_view["dex_path"].iloc[j]),
                            "dex_l1_match": str(df_view.get("DEX_L1", pd.Series([""])).iloc[i]) == str(df_view.get("DEX_L1", pd.Series([""])).iloc[j]),
                        }
                    )
        metric_D = pd.DataFrame(d_rows)
        total_pairs = int(len(metric_D))
        blocks_to_append["metric_D_path_stability"] = [
            {
                "title": "Summary (model | total_similar_pairs | pct_full_path_match | pct_l1_match | threshold_used)",
                "headers": ["model", "total_similar_pairs", "pct_full_path_match", "pct_l1_match", "threshold_used"],
                "rows": [
                    ["WALLE", total_pairs, float(metric_D["walle_path_match"].mean() * 100.0) if total_pairs else None, float(metric_D["walle_l1_match"].mean() * 100.0) if total_pairs else None, THRESHOLD],
                    ["DEX", total_pairs, float(metric_D["dex_path_match"].mean() * 100.0) if total_pairs else None, float(metric_D["dex_l1_match"].mean() * 100.0) if total_pairs else None, THRESHOLD],
                ],
            }
        ]
    except Exception as e:
        metric_D = pd.DataFrame([{"status": "skipped_missing_dependency_or_error", "reason": str(e)[:300]}])

    # ---- Metric F: resolution coherence ----
    f_rows: list[dict] = []
    for r in df_view.itertuples(index=False):
        inc_id = getattr(r, "INC_ID", None)
        wpath = getattr(r, "walle_path", None)
        dpath = getattr(r, "dex_path", None)
        res = getattr(r, "INC_RESOLUTION", None) if "INC_RESOLUTION" in df_view.columns else None
        work = getattr(r, "INC_COMMENTS", None) if "INC_COMMENTS" in df_view.columns else None

        w_full = _path_res_sim(wpath, res)
        d_full = _path_res_sim(dpath, res)
        w_l1l2 = _path_res_sim(_build_path(getattr(r, "WALLE_L1", None), getattr(r, "WALLE_L2", None), None, None, generic), res)
        d_l1l2 = _path_res_sim(_build_path(getattr(r, "DEX_L1", None), getattr(r, "DEX_L2", None), None, None, generic), res)
        w_work = _path_res_sim(wpath, work)
        d_work = _path_res_sim(dpath, work)

        def _winner(a, b) -> str:
            if a is None or b is None:
                return "Tie"
            if a > b + 0.05:
                return "WALLE"
            if b > a + 0.05:
                return "DEX"
            return "Tie"

        f_rows.append(
            {
                "INC_ID": inc_id,
                "walle_path": wpath,
                "dex_path": dpath,
                "walle_full_sim": w_full,
                "dex_full_sim": d_full,
                "walle_l1l2_sim": w_l1l2,
                "dex_l1l2_sim": d_l1l2,
                "walle_worknotes_sim": w_work,
                "dex_worknotes_sim": d_work,
                "resolution_available": bool(res) and str(res).strip() != "",
                "full_sim_winner": _winner(w_full, d_full),
                "l1l2_sim_winner": _winner(w_l1l2, d_l1l2),
            }
        )
    metric_F = pd.DataFrame(f_rows)
    scoreable = metric_F[metric_F["resolution_available"] == True]  # noqa: E712
    metric_F_summary = pd.DataFrame(
        [
            {
                "model": "WALLE",
                "scoreable_rows": int(len(scoreable)),
                "mean_full_sim": float(pd.to_numeric(scoreable["walle_full_sim"], errors="coerce").mean()) if len(scoreable) else None,
                "mean_l1l2_sim": float(pd.to_numeric(scoreable["walle_l1l2_sim"], errors="coerce").mean()) if len(scoreable) else None,
                "mean_worknotes_sim": float(pd.to_numeric(scoreable["walle_worknotes_sim"], errors="coerce").mean()) if len(scoreable) else None,
                "pct_winner_full": float((scoreable["full_sim_winner"] == "WALLE").mean() * 100.0) if len(scoreable) else None,
                "pct_winner_l1l2": float((scoreable["l1l2_sim_winner"] == "WALLE").mean() * 100.0) if len(scoreable) else None,
            },
            {
                "model": "DEX",
                "scoreable_rows": int(len(scoreable)),
                "mean_full_sim": float(pd.to_numeric(scoreable["dex_full_sim"], errors="coerce").mean()) if len(scoreable) else None,
                "mean_l1l2_sim": float(pd.to_numeric(scoreable["dex_l1l2_sim"], errors="coerce").mean()) if len(scoreable) else None,
                "mean_worknotes_sim": float(pd.to_numeric(scoreable["dex_worknotes_sim"], errors="coerce").mean()) if len(scoreable) else None,
                "pct_winner_full": float((scoreable["full_sim_winner"] == "DEX").mean() * 100.0) if len(scoreable) else None,
                "pct_winner_l1l2": float((scoreable["l1l2_sim_winner"] == "DEX").mean() * 100.0) if len(scoreable) else None,
            },
        ]
    )

    # ---- Metric E: confidence calibration (WALLE only; requires Metric 6 outputs) ----
    try:
        from scipy.stats import pearsonr

        conf_col = "WALLE_AI_L4_CONFIDENCE" if "WALLE_AI_L4_CONFIDENCE" in df_view.columns else None
        actionable_col = "WALLE_L4_ACTIONABLE" if "WALLE_L4_ACTIONABLE" in df_view.columns else None
        if conf_col is None:
            metric_E = pd.DataFrame([{"status": "skipped", "reason": "WALLE_AI_L4_CONFIDENCE not present in data sheet"}])
        else:
            j = metric_6_rows.copy()
            if "error" in j.columns:
                j = j[j["error"].isna()]
            keep = [c for c in ["INC_ID", "walle_overall_avg", "walle_section1_avg", "walle_section2_avg"] if c in j.columns]
            j = j[keep]
            m = df_view[["INC_ID", "walle_path", conf_col] + ([actionable_col] if actionable_col else [])].merge(j, on="INC_ID", how="inner")
            m["walle_confidence"] = pd.to_numeric(m[conf_col], errors="coerce")
            m["walle_judge_overall"] = pd.to_numeric(m.get("walle_overall_avg"), errors="coerce")
            m["walle_judge_section1"] = pd.to_numeric(m.get("walle_section1_avg"), errors="coerce")
            m["walle_judge_section2"] = pd.to_numeric(m.get("walle_section2_avg"), errors="coerce")
            m = m.dropna(subset=["walle_confidence", "walle_judge_overall"])
            if m.empty:
                metric_E = pd.DataFrame([{"status": "skipped", "reason": "No rows after join/dropna"}])
            else:
                r, pval = pearsonr(m["walle_confidence"], m["walle_judge_overall"])
                is_calibrated = bool((r > 0.3) and (pval < 0.05))
                bins = [0, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0]
                labels = ["<0.50", "0.50-0.60", "0.60-0.70", "0.70-0.80", "0.80-0.90", "0.90-1.00"]
                m["confidence_bin"] = pd.cut(m["walle_confidence"].clip(lower=0, upper=1.0), bins=bins, labels=labels, include_lowest=True)
                metric_E = m.rename(columns={actionable_col: "walle_actionable_flag"} if actionable_col else {})[
                    ["INC_ID", "walle_path", "walle_confidence"]
                    + (["walle_actionable_flag"] if actionable_col else [])
                    + ["walle_judge_overall", "walle_judge_section1", "walle_judge_section2", "confidence_bin"]
                ]

                blocks_to_append["metric_E_confidence_calibration"] = [
                    {"title": "pearson_r | pearson_p | is_calibrated", "headers": ["pearson_r", "pearson_p", "is_calibrated"], "rows": [[float(r), float(pval), is_calibrated]]}
                ]
                if "walle_actionable_flag" in metric_E.columns:
                    t = metric_E[metric_E["walle_actionable_flag"].apply(_is_trueish)]
                    f = metric_E[~metric_E["walle_actionable_flag"].apply(_is_trueish)]
                    mt = float(pd.to_numeric(t["walle_judge_overall"], errors="coerce").mean()) if not t.empty else None
                    mf = float(pd.to_numeric(f["walle_judge_overall"], errors="coerce").mean()) if not f.empty else None
                    valid = (mt is not None and mf is not None and abs(mt - mf) > 0.3)
                    blocks_to_append["metric_E_confidence_calibration"].append(
                        {"title": "Actionability", "headers": ["mean_judge_actionable_true", "mean_judge_actionable_false", "actionability_flag_valid"], "rows": [[mt, mf, bool(valid)]]}
                    )
                bstats = metric_E.groupby("confidence_bin", dropna=False)["walle_judge_overall"].agg(["count", "mean", "std"]).reset_index()
                blocks_to_append["metric_E_confidence_calibration"].append(
                    {
                        "title": "Bin stats",
                        "headers": ["confidence_bin", "count", "mean_judge_score", "std"],
                        "rows": [[str(rec["confidence_bin"]), int(rec["count"]), float(rec["mean"]) if pd.notna(rec["mean"]) else None, float(rec["std"]) if pd.notna(rec["std"]) else None] for rec in bstats.to_dict(orient="records")],
                    }
                )
    except Exception as e:
        metric_E = pd.DataFrame([{"status": "skipped_missing_dependency_or_error", "reason": str(e)[:300]}])

    # ---- Metric A/B/G: LLM-based (requires Azure env configured; reuse judge parallel patterns) ----
    import os
    azure_ready = bool(os.getenv("AZURE_OPENAI_API_KEY")) and bool(os.getenv("AZURE_OPENAI_ENDPOINT")) and bool(os.getenv("AZURE_OPENAI_DEPLOYMENT_NAME"))
    if bool(args.llm_judge) and azure_ready:
        # For brevity, reuse the existing judge agent stack via pydantic-ai Agents per metric.
        from pydantic import BaseModel, Field
        from pydantic_ai import Agent
        from pydantic_ai.models.openai import OpenAIChatModel, OpenAIChatModelSettings
        import asyncio

        deployment = (os.getenv("AZURE_OPENAI_DEPLOYMENT_NAME") or "").strip()
        model = OpenAIChatModel(deployment, provider="azure")
        x_upstream_env = (os.getenv("X_UPSTREAM_ENV") or os.getenv("X-Upstream-Env") or "").strip()
        project_id = (os.getenv("PROJECT_ID") or os.getenv("projectId") or "").strip()
        extra_headers: dict[str, str] = {}
        if x_upstream_env:
            extra_headers.update({"X-Upstream-Env": x_upstream_env, "projectId": project_id, "X-Model-Usage-Type": x_upstream_env, "modelUsageType": x_upstream_env})
        elif project_id:
            extra_headers["projectId"] = project_id
        ms = OpenAIChatModelSettings(extra_headers=extra_headers) if extra_headers else OpenAIChatModelSettings()

        # Metric A
        class AOut(BaseModel):
            l1l2: int = Field(..., ge=0, le=3)
            l2l3: int = Field(..., ge=0, le=3)
            l3l4: int = Field(..., ge=0, le=3)
            overall: int = Field(..., ge=0, le=3)
            weakest: str
            reason: str

        agentA = Agent(model, output_type=AOut, model_settings=ms)

        def _promptA(path: str) -> str:
            return (
                "You are evaluating an IT incident classification path for internal coherence.\n"
                "Score only structural consistency — not whether the classification is correct.\n\n"
                f"Path: {path}\n\n"
                "Score each level transition (0=contradiction, 1=unrelated, 2=loosely consistent, 3=fully consistent):\n"
                "  L1→L2: Does Category follow from Domain?\n"
                "  L2→L3: Does Subcategory follow from Category?\n"
                "  L3→L4: Does Key Issue follow from Subcategory?\n"
                "  overall: Is this a coherent business address for one incident type?\n\n"
                "Return ONLY valid JSON:\n"
                '{"l1l2":0-3,"l2l3":0-3,"l3l4":0-3,"overall":0-3,\n'
                ' "weakest":"L1→L2|L2→L3|L3→L4|None","reason":"one sentence"}\n'
            )
        A_base = df_view[["INC_ID", "walle_path", "dex_path"]].copy()
        A_base = A_base[(A_base["walle_path"].notna()) | (A_base["dex_path"].notna())].reset_index(drop=True)
        pw = [_promptA(p) for p in A_base["walle_path"].fillna("(none)").tolist()]
        pdx = [_promptA(p) for p in A_base["dex_path"].fillna("(none)").tolist()]
        ow = asyncio.run(_run_agent_prompts_parallel(agent=agentA, prompts=pw, workers=max(1, int(args.llm_judge_workers)), max_retries=int(args.llm_judge_max_retries), progress_label="metric_A_WALLE"))
        od = asyncio.run(_run_agent_prompts_parallel(agent=agentA, prompts=pdx, workers=max(1, int(args.llm_judge_workers)), max_retries=int(args.llm_judge_max_retries), progress_label="metric_A_DEX"))
        A_rows: list[dict] = []
        for i, rr in enumerate(A_base.itertuples(index=False)):
            row = {"INC_ID": rr.INC_ID, "walle_path": rr.walle_path, "dex_path": rr.dex_path}
            for obj, pref in [(ow[i], "walle"), (od[i], "dex")]:
                if obj.get("error"):
                    row[f"{pref}_error"] = obj.get("error")
                else:
                    row[f"{pref}_l1l2"] = obj.get("l1l2")
                    row[f"{pref}_l2l3"] = obj.get("l2l3")
                    row[f"{pref}_l3l4"] = obj.get("l3l4")
                    row[f"{pref}_overall"] = obj.get("overall")
                    row[f"{pref}_weakest"] = obj.get("weakest")
                    row[f"{pref}_reason"] = obj.get("reason")
            row["status"] = "success" if ("walle_error" not in row and "dex_error" not in row) else "llm_failed"
            A_rows.append(row)
        metric_A = pd.DataFrame(A_rows)
        metric_A_summary = pd.DataFrame(
            [
                {
                    "model": "WALLE",
                    "mean_l1l2": float(pd.to_numeric(metric_A.loc[metric_A["status"] == "success", "walle_l1l2"], errors="coerce").mean()) if not metric_A.empty else None,
                    "mean_l2l3": float(pd.to_numeric(metric_A.loc[metric_A["status"] == "success", "walle_l2l3"], errors="coerce").mean()) if not metric_A.empty else None,
                    "mean_l3l4": float(pd.to_numeric(metric_A.loc[metric_A["status"] == "success", "walle_l3l4"], errors="coerce").mean()) if not metric_A.empty else None,
                    "mean_overall": float(pd.to_numeric(metric_A.loc[metric_A["status"] == "success", "walle_overall"], errors="coerce").mean()) if not metric_A.empty else None,
                    "pct_score3": float((pd.to_numeric(metric_A.loc[metric_A["status"] == "success", "walle_overall"], errors="coerce") == 3).mean() * 100.0) if not metric_A.empty else None,
                    "pct_has_contradiction": float((metric_A.loc[metric_A["status"] == "success", ["walle_l1l2", "walle_l2l3", "walle_l3l4", "walle_overall"]].apply(pd.to_numeric, errors="coerce") == 0).any(axis=1).mean() * 100.0) if not metric_A.empty else None,
                    "most_common_weakest": metric_A.loc[metric_A["status"] == "success", "walle_weakest"].astype(str).value_counts().head(1).index[0] if "walle_weakest" in metric_A.columns and not metric_A.empty else None,
                },
                {
                    "model": "DEX",
                    "mean_l1l2": float(pd.to_numeric(metric_A.loc[metric_A["status"] == "success", "dex_l1l2"], errors="coerce").mean()) if not metric_A.empty else None,
                    "mean_l2l3": float(pd.to_numeric(metric_A.loc[metric_A["status"] == "success", "dex_l2l3"], errors="coerce").mean()) if not metric_A.empty else None,
                    "mean_l3l4": float(pd.to_numeric(metric_A.loc[metric_A["status"] == "success", "dex_l3l4"], errors="coerce").mean()) if not metric_A.empty else None,
                    "mean_overall": float(pd.to_numeric(metric_A.loc[metric_A["status"] == "success", "dex_overall"], errors="coerce").mean()) if not metric_A.empty else None,
                    "pct_score3": float((pd.to_numeric(metric_A.loc[metric_A["status"] == "success", "dex_overall"], errors="coerce") == 3).mean() * 100.0) if not metric_A.empty else None,
                    "pct_has_contradiction": float((metric_A.loc[metric_A["status"] == "success", ["dex_l1l2", "dex_l2l3", "dex_l3l4", "dex_overall"]].apply(pd.to_numeric, errors="coerce") == 0).any(axis=1).mean() * 100.0) if not metric_A.empty else None,
                    "most_common_weakest": metric_A.loc[metric_A["status"] == "success", "dex_weakest"].astype(str).value_counts().head(1).index[0] if "dex_weakest" in metric_A.columns and not metric_A.empty else None,
                },
            ]
        )

        # Metric B
        class BOut(BaseModel):
            action: int = Field(..., ge=0, le=2)
            action_reason: str
            system: int = Field(..., ge=0, le=2)
            system_reason: str
            routing: int = Field(..., ge=0, le=2)
            routing_reason: str
            composite: int
            automation_action: str | None = None

        agentB = Agent(model, output_type=BOut, model_settings=ms)

        def _promptB(path: str) -> str:
            return (
            "You are evaluating whether an IT incident classification path contains \n"
            "enough specificity to trigger or inform an automated resolution workflow.\n\n"
            f"Path: {path}\n\n"
            "Score three dimensions (0, 1, or 2):\n\n"
            "ACTION SPECIFICITY — Does the path indicate what action to take?\n"
            "  2=specific action stated (unlock account, clear cache, power cycle, redeploy)\n"
            "  1=action type implied but not specific (configuration change, access grant)\n"
            "  0=only the symptom described, no action signal\n\n"
            "SYSTEM SPECIFICITY — Does the path name a specific product or system?\n"
            "  2=specific named product (Cisco Secure Client, Windows Hello, Citrix Workspace)\n"
            "  1=product category (VPN client, authenticator app, business application)\n"
            "  0=generic type only (Software, Network, Application, Error/Failure)\n\n"
            "ROUTING UNIQUENESS — Could this path route to exactly one team or script?\n"
            "  2=uniquely routes to one team and one workflow\n"
            "  1=narrows to 2-3 teams or workflows\n"
            "  0=too broad, many teams could apply\n\n"
            "Return ONLY valid JSON:\n"
            '{"action":0-2,"action_reason":"one sentence",\n'
            ' "system":0-2,"system_reason":"one sentence",\n'
            ' "routing":0-2,"routing_reason":"one sentence",\n'
            ' "composite":sum_of_three,\n'
            ' "automation_action":"what specific action this enables or None"}\n'
            )

        bw = [_promptB(p) for p in A_base["walle_path"].fillna("(none)").tolist()]
        bd = [_promptB(p) for p in A_base["dex_path"].fillna("(none)").tolist()]
        obw = asyncio.run(_run_agent_prompts_parallel(agent=agentB, prompts=bw, workers=max(1, int(args.llm_judge_workers)), max_retries=int(args.llm_judge_max_retries), progress_label="metric_B_WALLE"))
        obd = asyncio.run(_run_agent_prompts_parallel(agent=agentB, prompts=bd, workers=max(1, int(args.llm_judge_workers)), max_retries=int(args.llm_judge_max_retries), progress_label="metric_B_DEX"))
        B_rows: list[dict] = []
        for i, rr in enumerate(A_base.itertuples(index=False)):
            row = {"INC_ID": rr.INC_ID, "walle_path": rr.walle_path, "dex_path": rr.dex_path}
            for obj, pref in [(obw[i], "walle"), (obd[i], "dex")]:
                if obj.get("error"):
                    row[f"{pref}_error"] = obj.get("error")
                else:
                    row[f"{pref}_action"] = obj.get("action")
                    row[f"{pref}_system"] = obj.get("system")
                    row[f"{pref}_routing"] = obj.get("routing")
                    row[f"{pref}_composite"] = obj.get("composite")
                    row[f"{pref}_automation_action"] = obj.get("automation_action")
            row["status"] = "success" if ("walle_error" not in row and "dex_error" not in row) else "llm_failed"
            B_rows.append(row)
        metric_B = pd.DataFrame(B_rows)
        metric_B_summary = pd.DataFrame(
            [
                {
                    "model": "WALLE",
                    "mean_composite": float(pd.to_numeric(metric_B.loc[metric_B["status"] == "success", "walle_composite"], errors="coerce").mean()) if not metric_B.empty else None,
                    "pct_high": float((pd.to_numeric(metric_B.loc[metric_B["status"] == "success", "walle_composite"], errors="coerce") >= 5).mean() * 100.0) if not metric_B.empty else None,
                    "pct_medium": float(((pd.to_numeric(metric_B.loc[metric_B["status"] == "success", "walle_composite"], errors="coerce") >= 3) & (pd.to_numeric(metric_B.loc[metric_B["status"] == "success", "walle_composite"], errors="coerce") <= 4)).mean() * 100.0) if not metric_B.empty else None,
                    "pct_low": float((pd.to_numeric(metric_B.loc[metric_B["status"] == "success", "walle_composite"], errors="coerce") <= 2).mean() * 100.0) if not metric_B.empty else None,
                    "mean_action": float(pd.to_numeric(metric_B.loc[metric_B["status"] == "success", "walle_action"], errors="coerce").mean()) if not metric_B.empty else None,
                    "mean_system": float(pd.to_numeric(metric_B.loc[metric_B["status"] == "success", "walle_system"], errors="coerce").mean()) if not metric_B.empty else None,
                    "mean_routing": float(pd.to_numeric(metric_B.loc[metric_B["status"] == "success", "walle_routing"], errors="coerce").mean()) if not metric_B.empty else None,
                },
                {
                    "model": "DEX",
                    "mean_composite": float(pd.to_numeric(metric_B.loc[metric_B["status"] == "success", "dex_composite"], errors="coerce").mean()) if not metric_B.empty else None,
                    "pct_high": float((pd.to_numeric(metric_B.loc[metric_B["status"] == "success", "dex_composite"], errors="coerce") >= 5).mean() * 100.0) if not metric_B.empty else None,
                    "pct_medium": float(((pd.to_numeric(metric_B.loc[metric_B["status"] == "success", "dex_composite"], errors="coerce") >= 3) & (pd.to_numeric(metric_B.loc[metric_B["status"] == "success", "dex_composite"], errors="coerce") <= 4)).mean() * 100.0) if not metric_B.empty else None,
                    "pct_low": float((pd.to_numeric(metric_B.loc[metric_B["status"] == "success", "dex_composite"], errors="coerce") <= 2).mean() * 100.0) if not metric_B.empty else None,
                    "mean_action": float(pd.to_numeric(metric_B.loc[metric_B["status"] == "success", "dex_action"], errors="coerce").mean()) if not metric_B.empty else None,
                    "mean_system": float(pd.to_numeric(metric_B.loc[metric_B["status"] == "success", "dex_system"], errors="coerce").mean()) if not metric_B.empty else None,
                    "mean_routing": float(pd.to_numeric(metric_B.loc[metric_B["status"] == "success", "dex_routing"], errors="coerce").mean()) if not metric_B.empty else None,
                },
            ]
        )

        # Metric G: cluster report from top 20 viable per model
        def classify_intervention(path: str) -> str:
            p = (path or "").lower()
            if any(w in p for w in ["account lock", "password reset", "unlock", "pin reset"]):
                return "Virtual Agent / Self-Service"
            if any(w in p for w in ["cache", "reboot", "restart", "power cycle"]):
                return "L1 Automation Script"
            if any(w in p for w in ["access denied", "provisioning", "license", "permission", "access - new"]):
                return "Access Request Automation"
            if any(w in p for w in ["citrix", "vpn", "zscaler", "wi-fi", "wifi", "network"]):
                return "Infrastructure Configuration Fix"
            if any(w in p for w in ["training", "how-to", "guidance", "informational"]):
                return "Knowledge Base Deflection"
            if any(w in p for w in ["outage", "isp", "vendor", "third-party", "third party"]):
                return "Vendor SLA / Escalation"
            return "Process Redesign"

        class GOut(BaseModel):
            intervention: str
            owning_team: str
            success_metric: str

        agentG = Agent(model, output_type=GOut, model_settings=ms)

        def _promptG(path: str, count: int, pct: float, itype: str) -> str:
            return (
            "You are an IT service improvement advisor helping leaders set incident \n"
            "reduction MBO targets.\n\n"
            f"Incident cluster path: {path}\n"
            f"Incidents in cluster: {count} ({pct}% of total)\n"
            f"Preliminary intervention type: {itype}\n\n"
            "Suggest a concrete plan for this cluster:\n"
            "1. One specific intervention to reduce these incidents (2-3 sentences, practical)\n"
            "2. Which team or role should own the reduction target\n"
            "3. How to measure success (one metric)\n\n"
            "Return ONLY valid JSON:\n"
            '{"intervention":"2-3 sentences","owning_team":"team name or role",\n'
            ' "success_metric":"how to measure"}\n'
            )
        viable = metric_C[metric_C["mbo_viable"] == True]  # noqa: E712
        topw = viable[viable["model"] == "WALLE"].head(20)
        topd = viable[viable["model"] == "DEX"].head(20)
        g_items: list[dict] = []
        g_prompts: list[str] = []
        for dfm, model_name in [(topw, "WALLE"), (topd, "DEX")]:
            for rr in dfm.itertuples(index=False):
                itype = classify_intervention(str(rr.full_path))
                g_items.append(
                    {
                        "model": model_name,
                        "rank": int(rr.rank_by_volume),
                        "full_path": rr.full_path,
                        "l1_domain": rr.l1_domain,
                        "l2_category": rr.l2_category,
                        "l3_subcategory": rr.l3_subcategory,
                        "l4_key_issue": rr.l4_key_issue,
                        "incident_count": int(rr.incident_count),
                        "pct_of_total": float(rr.pct_of_total),
                        "cumulative_pct": float(rr.cumulative_pct),
                        "intervention_type": itype,
                        "mean_path_consistency_score": None,
                        "mean_automation_score": None,
                    }
                )
                g_prompts.append(_promptG(str(rr.full_path), int(rr.incident_count), float(rr.pct_of_total), itype))
        og = asyncio.run(_run_agent_prompts_parallel(agent=agentG, prompts=g_prompts, workers=max(1, min(int(args.llm_judge_workers), 5)), max_retries=int(args.llm_judge_max_retries), progress_label="metric_G"))
        g_rows: list[dict] = []
        for i, base in enumerate(g_items):
            obj = og[i]
            row = dict(base)
            if obj.get("error"):
                row["status"] = "llm_failed"
            else:
                row["specific_intervention"] = obj.get("intervention")
                row["owning_team"] = obj.get("owning_team")
                row["success_metric"] = obj.get("success_metric")
                row["status"] = "success"
            g_rows.append(row)
        metric_G = pd.DataFrame(g_rows)
    else:
        metric_A = pd.DataFrame([{"status": "skipped", "reason": "Enable --llm-judge (and Azure env vars) for Metric A"}])
        metric_A_summary = pd.DataFrame([{"status": "skipped"}])
        metric_B = pd.DataFrame([{"status": "skipped", "reason": "Enable --llm-judge (and Azure env vars) for Metric B"}])
        metric_B_summary = pd.DataFrame([{"status": "skipped"}])
        metric_G = pd.DataFrame([{"status": "skipped", "reason": "Enable --llm-judge (and Azure env vars) for Metric G"}])

    per_sheet_lines: dict[str, list[str]] = {
        "data": [
            "HOW TO READ THIS SHEET (data)",
            "- This is the row-level joined dataset used by all metrics.",
            "- Use it for spot checks: filter where DEX_Lk != WALLE_Lk, then read INC_* text and WALLE_AI_* rationale/confidence.",
            f"- {usable_def}",
            "PITFALLS",
            "- Free text is noisy; disagreements can be taxonomy differences rather than model errors.",
        ],
        "metric_1_coverage": [
            "METRIC 1: COVERAGE / COMPLETENESS (metric_1_coverage)",
            "- Meaning: % of rows with a usable label at each level for DEX and WALLE.",
            "- Columns: level, rows, dex_coverage_pct, walle_coverage_pct.",
            f"- {usable_def}",
            "INFERENCES FROM THIS RUN (DEX vs WALLE)",
            *(cov_lines if cov_lines else ["- (no coverage comparisons available)"]),
            "HOW TO USE IT",
            "- If coverage drops sharply at L3/L4 for a model, downstream analytics/actionability at that level will suffer.",
            "PITFALLS",
            "- Higher coverage can be achieved by using broad catch-alls; validate with Metric 2.",
        ],
        "metric_2_granularity": [
            "METRIC 2: GRANULARITY / CONCENTRATION / ENTROPY (metric_2_granularity)",
            "- Meaning (usable labels only):",
            "  - unique_labels: number of distinct labels used (after normalization).",
            "  - top10_share_pct: % mass held by top 10 labels (higher = more concentrated).",
            "  - entropy_bits: spread of label distribution (higher = more diverse).",
            f"- {usable_def}",
            "INFERENCES FROM THIS RUN (DEX vs WALLE)",
            *(gran_lines if gran_lines else ["- (no granularity comparisons available)"]),
            "HOW TO USE IT",
            "- High top10_share + low entropy suggests a few dominant buckets (coarse taxonomy or weak differentiation).",
            "PITFALLS",
            "- More granularity is not more correctness. Validate via Metric 3 + spot checks in data.",
        ],
        "metric_2_top_labels": [
            "METRIC 2 (DETAIL): TOP LABELS (metric_2_top_labels)",
            "- Meaning: top labels per (model, level) with count and pct_of_usable.",
            "- Use it to understand what drives concentration seen in metric_2_granularity.",
            f"- {usable_def}",
            "HOW TO USE IT",
            "- If a small set of labels dominates L4, review those incidents to see if taxonomy is collapsing into a few buckets.",
            "PITFALLS",
            "- Labels are normalized (lower/strip). Formatting-only differences are intentionally collapsed.",
        ],
        "metric_3_agreement": [
            "METRIC 3: AGREEMENT (EXACT STRING MATCH) (metric_3_agreement)",
            "- Meaning: exact-match % between DEX and WALLE, computed only on rows where BOTH sides have usable labels.",
            "- Columns: level, rows_total, rows_both_usable, exact_match_pct.",
            f"- {usable_def}",
            "INFERENCES FROM THIS RUN (DEX vs WALLE)",
            *(agree_lines if agree_lines else ["- (no agreement comparisons available)"]),
            "HOW TO USE IT",
            "- Low agreement at L1/L2 suggests fundamental taxonomy mismatch; low mainly at L4 suggests different granularity/issue framing.",
            "PITFALLS",
            "- This is NOT semantic agreement; synonyms or near-matches count as disagreement.",
        ],
        "metric_3_disagreements": [
            "METRIC 3 (DETAIL): TOP DISAGREEMENT PAIRS (metric_3_disagreements)",
            "- Meaning: most frequent (DEX label -> WALLE label) mismatches among mutually usable rows.",
            "- Columns: level, dex_label, walle_label, count, pct_of_both_usable.",
            "HOW TO USE IT",
            "- High-frequency pairs usually indicate systematic mapping differences; candidates for taxonomy reconciliation/mapping tables.",
            "PITFALLS",
            "- Frequency does not imply which model is correct; validate by reading incidents in the data sheet.",
        ],
        "metric_4_walle_explainability": [
            "METRIC 4: WALLE EXPLAINABILITY / ACTIONABILITY QUALITY (metric_4_walle_explainability)",
            "- Meaning: presence/completeness of WALLE fields (rationale, keywords, root cause, confidence, resolution action, etc).",
            "- Slices include: all_rows, walle_l4_usable, and walle_l4_usable split by actionable True/False.",
            "- Column pct_present is computed within each slice; 'value' is used for confidence mean/median rows.",
            f"- {usable_def}",
            "INFERENCES FROM THIS RUN",
            *(m4_lines if m4_lines else ["- (insufficient columns or no rows to compute Metric 4)"]),
            "HOW TO USE IT",
            "- If WALLE is used operationally, missing confidence/rationale/actionability_reason reduces trust and usability even if labels exist.",
            "PITFALLS",
            "- This does not measure whether the rationale is correct—only that it is present.",
        ],
        "metric_4_confidence_bins": [
            "METRIC 4 (DETAIL): WALLE L4 CONFIDENCE DISTRIBUTION (metric_4_confidence_bins)",
            "- Meaning: histogram of numeric L4 confidence values (only where WALLE_L4 is usable).",
            "- Use it to understand whether the model is mostly low/medium/high confidence in the sampled dataset.",
            "PITFALLS",
            "- Confidence scale must be consistent (expected 0..1). Values outside range are shown in <0 or >1 bins.",
        ],
        "metric_5_stratified_agreement": [
            "METRIC 5: STRATIFIED AGREEMENT (metric_5_stratified_agreement)",
            "- Meaning: Metric 3 agreement recomputed within strata to localize where mismatch concentrates.",
            "- Strata computed: WALLE_L1 x month_bucket, and WALLE_L1/L2/L3 combinations (requires enough rows).",
            "- Only groups with rows_both_usable >= 20 are included (to avoid unstable percentages).",
            f"- {usable_def}",
            "INFERENCES FROM THIS RUN (focus on low agreement strata)",
            *(m5_lines if m5_lines else ["- (no strata met minimum rows / missing dates / missing WALLE hierarchy columns)"]),
            "HOW TO USE IT",
            "- Use low-agreement strata as the starting point for qualitative review and taxonomy mapping work.",
            "PITFALLS",
            "- Agreement is string match; taxonomy synonyms will show as disagreement.",
        ],
        "metric_5_top_disagreement_in_strata": [
            "METRIC 5 (DETAIL): TOP DISAGREEMENT PAIR PER STRATUM (metric_5_top_disagreement_in_strata)",
            "- Meaning: for each stratum+level, the single most common (DEX->WALLE) mismatch pair, to quickly explain low agreement.",
            "HOW TO USE IT",
            "- If the top pair repeats across many strata, it’s a strong candidate for a mapping table.",
            "PITFALLS",
            "- This reports only the top pair (not top-N); use metric_3_disagreements for global top pairs.",
        ],
        "metric_6_llm_judge_rows": [
            "METRIC 6: LLM-AS-JUDGE (ROW-LEVEL) (metric_6_llm_judge_rows)",
            "- Meaning: Full rubric — WALLE vs DEX on 7 scored dimensions (D1-D7), per-model section averages, and a single-incident verdict.",
            "- Scoring scale: 4 fully meets, 3 mostly, 2 partial, 1 weak, 0 contradicts, N not applicable.",
            "- Execution: incidents are judged in parallel (asyncio.Semaphore + gather) like L4 batch workers; row order in this sheet matches the sampled input order.",
            "- Incident text: INC_BRIEF_DESCRIPTION (+ error msg), long description bundle (ACTION/COMMENTS/UPDATE_ACTION), RESOLUTION, work notes (COMMENTS/MONITORING), steps (UPDATE_ACTION_ESS/UPDATE_ACTION). KB fields: not in extract → judge marks D3 as N.",
            "- Columns:",
            "  - Inputs: WALLE_L1..L4 and DEX_L1..L4",
            "  - Dimension outputs: walle_D1..D7_score + walle_D1..D7_reasoning (and same for dex_*)",
            "  - Examples: walle_D5_example_automation, walle_D6_example_process_improvement (and dex_*)",
            "  - Summary: walle_section1_avg (D1-D4), walle_section2_avg (D5-D7), walle_overall_avg (equal weight over applicable dims); and dex_* equivalents",
            "  - Verdict: verdict_overall, verdict_accuracy_leader, verdict_actionability_leader, verdict_dimension_largest_gap, verdict_one_sentence_summary",
            "INFERENCES FROM THIS RUN",
            *(llm_lines if llm_lines else ["- (LLM judge not run; enable with --llm-judge and set Azure OpenAI env vars)"]),
            "PITFALLS",
            "- Model-based proxy judge, not ground truth. Verdict applies to this row only; roll up conclusions manually across incidents.",
            "- If many rows have 'error' populated, reduce llm_judge_workers and/or judge_n; transient 429/5xx are retried but persistent errors will remain.",
            "PROMPT USED",
            "- Uses the full WALLE vs DEX 7-dimension rubric prompt defined in `_llm_judge_user_prompt(...)` in this script.",
        ],
        "metric_6_llm_judge_summary": [
            "METRIC 6: LLM-AS-JUDGE (SUMMARY) (metric_6_llm_judge_summary)",
            "- Meaning: Across SUCCESSFUL judge rows only (rows without error) — aggregates of verdict fields and mean reported averages.",
            "- How to read:",
            "  - metric=verdict_overall: % of rows where the judge picked WALLE vs DEX vs Tie (Tie if overall averages within 0.25 per rubric).",
            "  - metric=verdict_accuracy_leader / verdict_actionability_leader: % splits for which model led (or Equal).",
            "  - metric starting with mean_*: arithmetic mean of the judge-reported averages (0-4 scale).",
            "  - metric=mean_walle_minus_dex_overall_avg: positive means WALLE scored higher on average.",
            "  - metric=verdict_dimension_largest_gap_top: top dimensions (D1-D7) that most frequently had the largest gap.",
            "PITFALLS",
            "- Averages are as reported by the judge in structured output; validate sampling and prompt version across runs.",
            "- If you change the prompt, do not compare these summary numbers across runs without re-judging the same sample.",
        ],
        "metric_A_path_consistency": [
            "METRIC A: PATH CONSISTENCY (metric_A_path_consistency)",
            "- LLM scores structural coherence of the 4-level path (not correctness).",
            "PROMPT USED (exact logic; {path} substituted per row)",
            "- You are evaluating an IT incident classification path for internal coherence. Score only structural consistency — not whether the classification is correct.",
            "- Score: l1l2, l2l3, l3l4, overall (0-3). Return JSON with weakest transition and one-sentence reason.",
            "HOW TO INFER OUTCOMES",
            "- Higher mean_overall (closer to 3) => paths are more internally coherent.",
            "- Higher pct_has_contradiction (any score=0) => taxonomy levels contradict each other more often (bad).",
        ],
        "metric_A_path_consistency_summary": [
            "METRIC A (SUMMARY): PATH CONSISTENCY SUMMARY (metric_A_path_consistency_summary)",
            "- Two rows: WALLE and DEX aggregate means and weakest-link stats.",
            "HOW TO INFER OUTCOMES",
            "- If WALLE mean_overall > DEX mean_overall, WALLE paths are structurally more coherent on average (per this judge).",
            "- most_common_weakest tells you which transition (L1→L2 vs L2→L3 vs L3→L4) is breaking most often.",
        ],
        "metric_B_path_automation": [
            "METRIC B: PATH AUTOMATION (metric_B_path_automation)",
            "- LLM scores action/system/routing specificity (0-6 composite).",
            "PROMPT USED (exact logic; {path} substituted per row)",
            "- Scores ACTION SPECIFICITY (0-2), SYSTEM SPECIFICITY (0-2), ROUTING UNIQUENESS (0-2), plus composite=sum and automation_action.",
            "HOW TO INFER OUTCOMES",
            "- Higher mean_composite and higher pct_high (>=5) => paths are more usable for automation/routing.",
            "- If action mean is low but system is high, the model names products but not what to do; if routing is low, labels are too broad for team/script selection.",
        ],
        "metric_B_path_automation_summary": [
            "METRIC B (SUMMARY): PATH AUTOMATION SUMMARY (metric_B_path_automation_summary)",
            "- Two rows: WALLE and DEX with composite distributions and dimension means.",
            "HOW TO INFER OUTCOMES",
            "- Compare mean_action/mean_system/mean_routing to see what drives composite differences.",
        ],
        "metric_C_mbo_clusters": [
            "METRIC C: MBO CLUSTERS (metric_C_mbo_clusters)",
            "- Group by full path to find Pareto clusters and mbo_viable targets.",
        ],
        "metric_D_path_stability": [
            "METRIC D: PATH STABILITY (metric_D_path_stability)",
            "- For text-similar pairs (TF-IDF cosine>=0.35), check whether each model assigns the same path and L1.",
        ],
        "metric_E_confidence_calibration": [
            "METRIC E: CONFIDENCE CALIBRATION (metric_E_confidence_calibration)",
            "- Join WALLE L4 confidence to Metric 6 judge scores; compute Pearson r and bin means.",
        ],
        "metric_F_resolution_coherence": [
            "METRIC F: RESOLUTION COHERENCE (metric_F_resolution_coherence)",
            "- TF-IDF similarity between model path and resolution/work-notes text; includes winner flags.",
        ],
        "metric_F_coherence_summary": [
            "METRIC F (SUMMARY): COHERENCE SUMMARY (metric_F_coherence_summary)",
            "- Two rows: WALLE and DEX mean similarities and win rates.",
        ],
        "metric_G_mbo_cluster_report": [
            "METRIC G: MBO CLUSTER REPORT (metric_G_mbo_cluster_report)",
            "- Top 20 mbo_viable clusters per model with rule-based intervention_type + LLM plan.",
            "PROMPT USED (per cluster)",
            "- Inputs: path, cluster count + pct_of_total, preliminary intervention_type. Output JSON: intervention, owning_team, success_metric.",
            "HOW TO INFER OUTCOMES",
            "- Prefer clusters with high incident_count/pct_of_total AND decent mean_path_consistency_score + mean_automation_score: these are high-leverage and well-defined targets.",
        ],
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as xw:
        df_view.to_excel(xw, sheet_name="data", index=False)
        metric_1.to_excel(xw, sheet_name="metric_1_coverage", index=False)
        metric_2.to_excel(xw, sheet_name="metric_2_granularity", index=False)
        metric_2_top.to_excel(xw, sheet_name="metric_2_top_labels", index=False)
        metric_3.to_excel(xw, sheet_name="metric_3_agreement", index=False)
        metric_3_disagree.to_excel(xw, sheet_name="metric_3_disagreements", index=False)
        metric_4.to_excel(xw, sheet_name="metric_4_walle_explainability", index=False)
        metric_4_conf.to_excel(xw, sheet_name="metric_4_confidence_bins", index=False)
        metric_5.to_excel(xw, sheet_name="metric_5_stratified_agreement", index=False)
        metric_5_pairs.to_excel(xw, sheet_name="metric_5_top_disagreement_in_strata", index=False)
        metric_6_rows.to_excel(xw, sheet_name="metric_6_llm_judge_rows", index=False)
        metric_6_summary.to_excel(xw, sheet_name="metric_6_llm_judge_summary", index=False)
        metric_A.to_excel(xw, sheet_name="metric_A_path_consistency", index=False)
        metric_A_summary.to_excel(xw, sheet_name="metric_A_path_consistency_summary", index=False)
        metric_B.to_excel(xw, sheet_name="metric_B_path_automation", index=False)
        metric_B_summary.to_excel(xw, sheet_name="metric_B_path_automation_summary", index=False)
        metric_C.to_excel(xw, sheet_name="metric_C_mbo_clusters", index=False)
        metric_D.to_excel(xw, sheet_name="metric_D_path_stability", index=False)
        metric_E.to_excel(xw, sheet_name="metric_E_confidence_calibration", index=False)
        metric_F.to_excel(xw, sheet_name="metric_F_resolution_coherence", index=False)
        metric_F_summary.to_excel(xw, sheet_name="metric_F_coherence_summary", index=False)
        metric_G.to_excel(xw, sheet_name="metric_G_mbo_cluster_report", index=False)

    # Prepend explainability into each sheet (so guidance lives with the tab).
    _prepend_sheet_explainability(output_path, per_sheet_lines)
    if blocks_to_append:
        _append_blocks_after_write(output_path, blocks_to_append)

    print(f"Wrote report: {output_path}", flush=True)
    print(metric_1.to_string(index=False), flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

